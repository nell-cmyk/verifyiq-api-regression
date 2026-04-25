#!/usr/bin/env python3
"""Generate the shared machine-readable fixture registry from curated sources.

Primary source (human-maintained):
  tools/fixture_registry_source/qa_fixture_registry.xlsx
Supplemental source (JSON-onboarded fixtures):
  tools/fixture_registry_source/supplemental_fixture_registry.yaml
GT extraction metadata source:
  tools/fixture_registry_source/gt_extraction_fixture_overrides.yaml
Outputs (automation source of truth plus /parse compatibility copy):
  tests/fixtures/fixture_registry.yaml
  tests/endpoints/parse/fixture_registry.yaml

Re-run this after editing the spreadsheet or supplemental YAML. Idempotent.

Usage:
  python3 -m venv .venv
  ./.venv/bin/python -m pip install -r tools/requirements.txt
  ./.venv/bin/python tools/generate_fixture_registry.py

Classification rules (schema_version: 2):
  - Rows with no gs:// path are dropped.
  - fileType == "No fileType" or "Fraud - Skipped"
      → verification_status=excluded, enabled=False (kept for traceability).
  - fileType containing "||" (composite label) is SPLIT into one record per
    individual fileType. Each split record is enabled on its own merit; the
    original composite label is preserved in `source_file_type` and the source
    row number in `source_row`.
  - fileType Status "✓"        → verification_status=confirmed, enabled=True.
  - fileType Status "⚠ Verify" → verification_status=unverified, enabled=True.
    ("Verify" is a human QA tag, not an automation gate — both are valid
    fixture candidates for Phase 2.)
  - anything else              → verification_status=unknown, enabled=False.
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from functools import lru_cache
from pathlib import Path
from typing import Any

import yaml

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from tests.endpoints.parse.fixture_json import unsupported_fixture_reason

SOURCE_XLSX = REPO_ROOT / "tools" / "fixture_registry_source" / "qa_fixture_registry.xlsx"
SUPPLEMENTAL_YAML = (
    REPO_ROOT / "tools" / "fixture_registry_source" / "supplemental_fixture_registry.yaml"
)
GT_EXTRACTION_OVERRIDES_YAML = (
    REPO_ROOT / "tools" / "fixture_registry_source" / "gt_extraction_fixture_overrides.yaml"
)
SHARED_OUTPUT_YAML = REPO_ROOT / "tests" / "fixtures" / "fixture_registry.yaml"
PARSE_COMPAT_OUTPUT_YAML = REPO_ROOT / "tests" / "endpoints" / "parse" / "fixture_registry.yaml"
OUTPUT_YAML = SHARED_OUTPUT_YAML
OUTPUT_PATHS = (SHARED_OUTPUT_YAML, PARSE_COMPAT_OUTPUT_YAML)
SHEET_NAME = "All"
HEADER_ROW = 4  # Rows 1–2 are title, row 3 blank, row 4 is the header. Data from row 5.
SCHEMA_VERSION = 2
SUPPLEMENTAL_SCHEMA_VERSION = 1
GT_EXTRACTION_OVERRIDES_SCHEMA_VERSION = 1

EXCLUDED_FILE_TYPES = {"No fileType", "Fraud - Skipped"}
COMPOSITE_DELIMITER = "||"
CONFIRMED_MARK = "✓"
UNVERIFIED_MARK = "⚠ Verify"
VALID_VERIFICATION_STATUSES = {"confirmed", "unverified", "excluded", "unknown"}
GT_EXTRACTION_STRING_KEYS = {
    "batch_expected_warning",
    "batch_expected_error_type",
    "batch_expected_error",
    "gt_extraction_skip_reason",
    "gt_extraction_classification",
    "gt_recovery_action",
}
GT_EXTRACTION_BOOL_KEYS = {
    "gt_extraction_eligible",
    "gt_clean_eligible",
    "negative_audit_useful",
}
VALID_GT_EXTRACTION_SKIP_REASONS = {
    "document_size_guard",
    "multi_account_document",
    "unsupported_fixture",
    "quality_gate_no_payload",
}
VALID_GT_EXTRACTION_CLASSIFICATIONS = {
    "fixture_too_large",
    "multi_account_fixture",
    "unsupported_artifact",
    "fixture_quality_gate_failed",
}
VALID_GT_RECOVERY_ACTIONS = {
    "replace_fixture",
    "split_fixture",
    "reduce_fixture",
    "review_fixture_quality",
    "keep_as_negative_coverage",
}


def build_parser() -> argparse.ArgumentParser:
    return argparse.ArgumentParser(
        description="Generate the shared machine-readable fixture registry from curated sources.",
        epilog=(
            "Bootstrap:\n"
            "  python3 -m venv .venv\n"
            "  ./.venv/bin/python -m pip install -r tools/requirements.txt\n\n"
            "Run:\n"
            "  ./.venv/bin/python tools/generate_fixture_registry.py"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )


def _display_path(path: Path) -> str:
    try:
        return path.resolve().relative_to(REPO_ROOT).as_posix()
    except ValueError:
        return str(path.expanduser().resolve())


def classify(file_type: str, file_type_status: str) -> tuple[str, bool]:
    """Return (verification_status, enabled) for a (post-split) fileType."""
    if file_type in EXCLUDED_FILE_TYPES:
        return "excluded", False
    if file_type_status == CONFIRMED_MARK:
        return "confirmed", True
    if file_type_status == UNVERIFIED_MARK:
        return "unverified", True
    return "unknown", False


def base_name_from_uri(uri: str) -> str:
    base = uri.rsplit("/", 1)[-1]
    stem = base.rsplit(".", 1)[0] if "." in base else base
    return stem.strip() or "fixture"


def reserve_name(candidate: str, used: set[str]) -> str:
    """Return a name unique within `used`, appending __2, __3, ... on collision."""
    name = candidate
    n = 2
    while name in used:
        name = f"{candidate}__{n}"
        n += 1
    used.add(name)
    return name


def default_enabled_for_status(verification_status: str) -> bool:
    return verification_status in {"confirmed", "unverified"}


def _validate_gt_metadata(
    metadata: dict[str, Any],
    *,
    context: str,
    override_yaml: Path,
) -> dict[str, Any]:
    allowed_keys = GT_EXTRACTION_STRING_KEYS | GT_EXTRACTION_BOOL_KEYS
    unknown = sorted(set(metadata) - allowed_keys)
    if unknown:
        raise RuntimeError(
            f"Invalid GT extraction overrides at {override_yaml}: "
            f"{context} has unknown metadata keys: {unknown}"
        )

    normalized: dict[str, Any] = {}
    for key in GT_EXTRACTION_STRING_KEYS:
        if key not in metadata:
            continue
        value = metadata[key]
        if not isinstance(value, str) or not value.strip():
            raise RuntimeError(
                f"Invalid GT extraction overrides at {override_yaml}: "
                f"{context} has invalid '{key}': {value!r}"
            )
        normalized[key] = value.strip()

    for key in GT_EXTRACTION_BOOL_KEYS:
        if key not in metadata:
            continue
        value = metadata[key]
        if not isinstance(value, bool):
            raise RuntimeError(
                f"Invalid GT extraction overrides at {override_yaml}: "
                f"{context} has invalid '{key}': {value!r}"
            )
        normalized[key] = value

    skip_reason = normalized.get("gt_extraction_skip_reason")
    if skip_reason is not None and skip_reason not in VALID_GT_EXTRACTION_SKIP_REASONS:
        raise RuntimeError(
            f"Invalid GT extraction overrides at {override_yaml}: "
            f"{context} has invalid 'gt_extraction_skip_reason': {skip_reason!r}"
        )

    classification = normalized.get("gt_extraction_classification")
    if classification is not None and classification not in VALID_GT_EXTRACTION_CLASSIFICATIONS:
        raise RuntimeError(
            f"Invalid GT extraction overrides at {override_yaml}: "
            f"{context} has invalid 'gt_extraction_classification': {classification!r}"
        )

    recovery_action = normalized.get("gt_recovery_action")
    if recovery_action is not None and recovery_action not in VALID_GT_RECOVERY_ACTIONS:
        raise RuntimeError(
            f"Invalid GT extraction overrides at {override_yaml}: "
            f"{context} has invalid 'gt_recovery_action': {recovery_action!r}"
        )

    if normalized.get("gt_extraction_eligible") is False:
        required = {
            "gt_extraction_skip_reason",
            "gt_extraction_classification",
            "gt_recovery_action",
        }
        missing = sorted(required - set(normalized))
        if missing:
            raise RuntimeError(
                f"Invalid GT extraction overrides at {override_yaml}: "
                f"{context} marks GT extraction ineligible but is missing {missing}"
            )

    return normalized


@lru_cache(maxsize=None)
def _load_gt_extraction_overrides_by_key(override_yaml_text: str) -> dict[tuple[str, str], dict[str, Any]]:
    override_yaml = Path(override_yaml_text).expanduser().resolve()
    if not override_yaml.exists():
        return {}

    try:
        doc = yaml.safe_load(override_yaml.read_text(encoding="utf-8"))
    except yaml.YAMLError as exc:
        raise RuntimeError(f"Invalid GT extraction overrides at {override_yaml}: {exc}") from exc

    if doc is None:
        return {}
    if not isinstance(doc, dict):
        raise RuntimeError(
            f"Invalid GT extraction overrides at {override_yaml}: "
            f"top-level document must be a mapping, got {type(doc).__name__}"
        )

    schema_version = doc.get("schema_version", GT_EXTRACTION_OVERRIDES_SCHEMA_VERSION)
    if not isinstance(schema_version, int) or isinstance(schema_version, bool):
        raise RuntimeError(
            f"Invalid GT extraction overrides at {override_yaml}: "
            f"'schema_version' must be an integer, got {schema_version!r}"
        )

    groups = doc.get("groups", [])
    if not isinstance(groups, list):
        raise RuntimeError(
            f"Invalid GT extraction overrides at {override_yaml}: "
            f"'groups' must be a list, got {type(groups).__name__}"
        )

    overrides: dict[tuple[str, str], dict[str, Any]] = {}
    for group_index, group in enumerate(groups, start=1):
        if not isinstance(group, dict):
            raise RuntimeError(
                f"Invalid GT extraction overrides at {override_yaml}: "
                f"group #{group_index} must be a mapping, got {type(group).__name__}"
            )
        group_metadata = group.get("metadata", {})
        if not isinstance(group_metadata, dict):
            raise RuntimeError(
                f"Invalid GT extraction overrides at {override_yaml}: "
                f"group #{group_index} metadata must be a mapping, got {type(group_metadata).__name__}"
            )
        normalized_group_metadata = _validate_gt_metadata(
            group_metadata,
            context=f"group #{group_index} metadata",
            override_yaml=override_yaml,
        )
        fixtures = group.get("fixtures", [])
        if not isinstance(fixtures, list):
            raise RuntimeError(
                f"Invalid GT extraction overrides at {override_yaml}: "
                f"group #{group_index} fixtures must be a list, got {type(fixtures).__name__}"
            )

        for fixture_index, fixture in enumerate(fixtures, start=1):
            if not isinstance(fixture, dict):
                raise RuntimeError(
                    f"Invalid GT extraction overrides at {override_yaml}: "
                    f"group #{group_index} fixture #{fixture_index} must be a mapping, "
                    f"got {type(fixture).__name__}"
                )
            gcs_uri = fixture.get("gcs_uri")
            file_type = fixture.get("file_type")
            if not isinstance(gcs_uri, str) or not gcs_uri.startswith("gs://"):
                raise RuntimeError(
                    f"Invalid GT extraction overrides at {override_yaml}: "
                    f"group #{group_index} fixture #{fixture_index} has invalid 'gcs_uri': {gcs_uri!r}"
                )
            if not isinstance(file_type, str) or not file_type.strip():
                raise RuntimeError(
                    f"Invalid GT extraction overrides at {override_yaml}: "
                    f"group #{group_index} fixture #{fixture_index} has invalid 'file_type': {file_type!r}"
                )

            fixture_metadata = {
                key: value
                for key, value in fixture.items()
                if key not in {"gcs_uri", "file_type"}
            }
            normalized_fixture_metadata = _validate_gt_metadata(
                fixture_metadata,
                context=f"group #{group_index} fixture #{fixture_index}",
                override_yaml=override_yaml,
            )
            key = (gcs_uri, file_type.strip())
            if key in overrides:
                raise RuntimeError(
                    f"Invalid GT extraction overrides at {override_yaml}: "
                    f"duplicate fixture override {key!r}"
                )
            overrides[key] = normalized_group_metadata | normalized_fixture_metadata

    return overrides


def fixture_metadata_overrides_for(
    *,
    gcs_uri: str,
    file_type: str | None,
    override_yaml: Path = GT_EXTRACTION_OVERRIDES_YAML,
) -> dict[str, Any]:
    overrides = _load_gt_extraction_overrides_by_key(str(Path(override_yaml).expanduser().resolve()))
    override = overrides.get((gcs_uri, file_type or ""))
    if override is None:
        return {}
    return dict(override)


def _gt_metadata_for_unsupported_fixture(unsupported_reason: str) -> dict[str, Any]:
    return {
        "fixture_unsupported_reason": unsupported_reason,
        "gt_extraction_eligible": False,
        "gt_extraction_skip_reason": "unsupported_fixture",
        "gt_extraction_classification": "unsupported_artifact",
        "gt_clean_eligible": False,
        "negative_audit_useful": True,
        "gt_recovery_action": "replace_fixture",
    }


def load_supplemental_registry_doc(
    supplemental_yaml: Path = SUPPLEMENTAL_YAML,
) -> dict:
    if not supplemental_yaml.exists():
        return {
            "schema_version": SUPPLEMENTAL_SCHEMA_VERSION,
            "fixtures": [],
        }

    try:
        doc = yaml.safe_load(supplemental_yaml.read_text(encoding="utf-8"))
    except yaml.YAMLError as exc:
        raise RuntimeError(
            f"Invalid supplemental fixture registry at {supplemental_yaml}: {exc}"
        ) from exc

    if doc is None:
        doc = {}
    if not isinstance(doc, dict):
        raise RuntimeError(
            f"Invalid supplemental fixture registry at {supplemental_yaml}: "
            f"top-level document must be a mapping, got {type(doc).__name__}"
        )

    schema_version = doc.get("schema_version", SUPPLEMENTAL_SCHEMA_VERSION)
    if not isinstance(schema_version, int) or isinstance(schema_version, bool):
        raise RuntimeError(
            f"Invalid supplemental fixture registry at {supplemental_yaml}: "
            f"'schema_version' must be an integer, got {schema_version!r}"
        )

    fixtures = doc.get("fixtures", [])
    if not isinstance(fixtures, list):
        raise RuntimeError(
            f"Invalid supplemental fixture registry at {supplemental_yaml}: "
            f"'fixtures' must be a list, got {type(fixtures).__name__}"
        )

    return {
        "schema_version": schema_version,
        "fixtures": fixtures,
    }


def _load_spreadsheet_fixtures(
    source_xlsx: Path = SOURCE_XLSX,
    gt_extraction_overrides_yaml: Path = GT_EXTRACTION_OVERRIDES_YAML,
) -> tuple[list[dict], Counter[str], int, set[str], set[tuple[str, str | None]]]:
    if not source_xlsx.exists():
        raise RuntimeError(f"source spreadsheet not found at {source_xlsx}")

    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "openpyxl is required for fixture-registry generation. "
            "Install tool deps with `./.venv/bin/python -m pip install -r tools/requirements.txt`."
        ) from exc

    wb = openpyxl.load_workbook(source_xlsx, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"sheet {SHEET_NAME!r} not found; sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    fixtures: list[dict] = []
    used_names: set[str] = set()
    counts: Counter[str] = Counter()
    existing_pairs: set[tuple[str, str | None]] = set()
    split_rows = 0

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True),
        start=HEADER_ROW + 1,
    ):
        folder, file_type, gcs_uri, ft_status, assignee, row_status, *_ = row
        gcs_uri = gcs_uri.strip() if isinstance(gcs_uri, str) else gcs_uri
        if not gcs_uri or not str(gcs_uri).startswith("gs://"):
            counts["skipped_no_gcs"] += 1
            continue

        raw_file_type = (file_type or "").strip()
        ft_status_clean = (ft_status or "").strip()
        folder_clean = (folder or "").strip() or None
        assignee_clean = (assignee or "").strip() or None
        row_status_clean = (row_status or "").strip() or None
        base_name = base_name_from_uri(str(gcs_uri))

        is_composite = COMPOSITE_DELIMITER in raw_file_type
        if is_composite:
            split_rows += 1
            parts = [p.strip() for p in raw_file_type.split(COMPOSITE_DELIMITER) if p.strip()]
        else:
            parts = [raw_file_type]

        for part in parts:
            verification_status, enabled = classify(part, ft_status_clean)
            counts[verification_status] += 1
            candidate = f"{base_name}__{part}" if is_composite and part else base_name
            metadata = fixture_metadata_overrides_for(
                gcs_uri=str(gcs_uri),
                file_type=part or None,
                override_yaml=gt_extraction_overrides_yaml,
            )
            unsupported_reason = unsupported_fixture_reason(str(gcs_uri))
            if unsupported_reason is not None:
                metadata |= _gt_metadata_for_unsupported_fixture(unsupported_reason)
            fixtures.append({
                "name": reserve_name(candidate, used_names),
                "file_type": part or None,
                "gcs_uri": str(gcs_uri),
                "source_folder": folder_clean,
                "source_file_type": raw_file_type or None,
                "source_file_type_status": ft_status_clean or None,
                "source_assignee": assignee_clean,
                "source_workflow_status": row_status_clean,
                "source_row": row_idx,
                "verification_status": verification_status,
                "enabled": enabled,
            } | metadata)
            existing_pairs.add((str(gcs_uri), part or None))

    return fixtures, counts, split_rows, used_names, existing_pairs


def _load_supplemental_fixtures(
    *,
    used_names: set[str],
    existing_pairs: set[tuple[str, str | None]],
    counts: Counter[str],
    supplemental_yaml: Path = SUPPLEMENTAL_YAML,
    gt_extraction_overrides_yaml: Path = GT_EXTRACTION_OVERRIDES_YAML,
) -> list[dict]:
    fixtures: list[dict] = []
    seen_pairs: set[tuple[str, str]] = set()

    for index, raw in enumerate(load_supplemental_registry_doc(supplemental_yaml)["fixtures"], start=1):
        if not isinstance(raw, dict):
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} must be a mapping, got {type(raw).__name__}"
            )

        gcs_uri = raw.get("gcs_uri")
        file_type = raw.get("file_type")
        if not isinstance(gcs_uri, str) or not gcs_uri.startswith("gs://"):
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} has invalid 'gcs_uri': {gcs_uri!r}"
            )
        unsupported_reason = unsupported_fixture_reason(gcs_uri)
        if unsupported_reason is not None:
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} has unsupported 'gcs_uri': {gcs_uri!r} ({unsupported_reason})"
            )
        if not isinstance(file_type, str) or not file_type.strip():
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} has invalid 'file_type': {file_type!r}"
            )

        key = (gcs_uri, file_type.strip())
        if key in seen_pairs:
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"duplicate supplemental fixture {key!r}"
            )
        seen_pairs.add(key)

        if key in existing_pairs:
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} duplicates an existing curated source fixture {key!r}"
            )

        verification_status = raw.get("verification_status", "unverified")
        if verification_status not in VALID_VERIFICATION_STATUSES:
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} has invalid 'verification_status': {verification_status!r}"
            )

        enabled = raw.get("enabled")
        if enabled is None:
            enabled = default_enabled_for_status(verification_status)
        if not isinstance(enabled, bool):
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} has invalid 'enabled': {enabled!r}"
            )

        source_folder = raw.get("source_folder") or file_type.strip()
        source_file_type = raw.get("source_file_type") or file_type.strip()
        name = raw.get("name") or base_name_from_uri(gcs_uri)
        if not isinstance(source_folder, str) or not source_folder.strip():
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} has invalid 'source_folder': {source_folder!r}"
            )
        if not isinstance(source_file_type, str) or not source_file_type.strip():
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} has invalid 'source_file_type': {source_file_type!r}"
            )
        if not isinstance(name, str) or not name.strip():
            raise RuntimeError(
                f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                f"fixture #{index} has invalid 'name': {name!r}"
            )
        source_file_type_status = raw.get("source_file_type_status")
        if source_file_type_status is None:
            source_file_type_status = {
                "confirmed": CONFIRMED_MARK,
                "unverified": UNVERIFIED_MARK,
            }.get(verification_status, "")
        source_assignee = raw.get("source_assignee")
        source_workflow_status = raw.get("source_workflow_status")
        for key, value in (
            ("source_file_type_status", source_file_type_status),
            ("source_assignee", source_assignee),
            ("source_workflow_status", source_workflow_status),
        ):
            if value is not None and not isinstance(value, str):
                raise RuntimeError(
                    f"Invalid supplemental fixture registry at {supplemental_yaml}: "
                    f"fixture #{index} has invalid '{key}': {value!r}"
                )

        counts[verification_status] += 1
        fixtures.append({
            "name": reserve_name(name.strip(), used_names),
            "file_type": file_type.strip(),
            "gcs_uri": gcs_uri,
            "source_folder": source_folder.strip(),
            "source_file_type": source_file_type.strip(),
            "source_file_type_status": source_file_type_status.strip() if source_file_type_status else None,
            "source_assignee": source_assignee.strip() if source_assignee else None,
            "source_workflow_status": source_workflow_status.strip() if source_workflow_status else None,
            "source_row": 0,
            "verification_status": verification_status,
            "enabled": enabled,
        } | fixture_metadata_overrides_for(
            gcs_uri=gcs_uri,
            file_type=file_type.strip(),
            override_yaml=gt_extraction_overrides_yaml,
        ))
        existing_pairs.add(key)

    return fixtures


def build_registry_document(
    *,
    source_xlsx: Path = SOURCE_XLSX,
    supplemental_yaml: Path | None = SUPPLEMENTAL_YAML,
    gt_extraction_overrides_yaml: Path = GT_EXTRACTION_OVERRIDES_YAML,
) -> dict:
    spreadsheet_fixtures, counts, split_rows, used_names, existing_pairs = _load_spreadsheet_fixtures(
        source_xlsx,
        gt_extraction_overrides_yaml=gt_extraction_overrides_yaml,
    )
    supplemental_fixtures = []
    if supplemental_yaml is not None:
        supplemental_fixtures = _load_supplemental_fixtures(
            used_names=used_names,
            existing_pairs=existing_pairs,
            counts=counts,
            supplemental_yaml=supplemental_yaml,
            gt_extraction_overrides_yaml=gt_extraction_overrides_yaml,
        )
    fixtures = spreadsheet_fixtures + supplemental_fixtures
    fixtures.sort(key=lambda f: (f["source_folder"] or "", f["file_type"] or "", f["name"]))

    doc = {
        "schema_version": SCHEMA_VERSION,
        "source": _display_path(source_xlsx),
        "total": len(fixtures),
        "composite_rows_split": split_rows,
        "counts": dict(sorted(counts.items())),
        "fixtures": fixtures,
    }
    if supplemental_fixtures:
        doc["supplemental_source"] = _display_path(supplemental_yaml or SUPPLEMENTAL_YAML)
    return doc


def write_registry_document(
    doc: dict,
    *,
    output_paths: tuple[Path, ...] = OUTPUT_PATHS,
) -> None:
    header = (
        "# AUTO-GENERATED by tools/generate_fixture_registry.py - do not edit by hand.\n"
        f"# Source: {doc.get('source', _display_path(SOURCE_XLSX))}\n"
        "# Edit the source registry files and rerun the generator to update this file.\n"
    )
    for output_yaml in output_paths:
        output_yaml.parent.mkdir(parents=True, exist_ok=True)
        with output_yaml.open("w", encoding="utf-8", newline="\n") as f:
            f.write(header)
            yaml.safe_dump(doc, f, allow_unicode=True, sort_keys=False, width=160)


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    parser.parse_args(argv)

    try:
        doc = build_registry_document()
        write_registry_document(doc)
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    enabled_count = sum(1 for x in doc["fixtures"] if x["enabled"])
    print(
        f"Wrote {len(doc['fixtures'])} fixture records ({enabled_count} enabled; "
        f"{doc['composite_rows_split']} source rows split into multiple records) to "
        + ", ".join(path.relative_to(REPO_ROOT).as_posix() for path in OUTPUT_PATHS),
        file=sys.stderr,
    )
    for k, v in sorted(doc["counts"].items()):
        print(f"  {k:28s} {v}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
