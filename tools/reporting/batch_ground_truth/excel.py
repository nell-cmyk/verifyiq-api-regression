from __future__ import annotations

import json
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import ExportRow, TemplateLayout
from .schema import FIXED_METADATA_HEADERS, build_main_sheet_header_order

JSON_WIDTH_HEADERS = {
    "summary_json",
    "raw_response_json",
    "raw_result_json",
    "transactions_json",
}

META_SHEET_TITLE = "_meta"


def _safe_sheet_title(file_type: str) -> str:
    cleaned = "".join("_" if ch in '[]:*?/\\' else ch for ch in file_type).strip()
    return (cleaned or "Sheet")[:31]


def _safe_filename(file_type: str) -> str:
    cleaned = "".join("_" if ch in '<>:"/\\|?*' else ch for ch in file_type).strip()
    return cleaned or "output"


def _column_width(header: str, layout: TemplateLayout) -> float:
    if header in layout.width_by_header and layout.width_by_header[header] is not None:
        return float(layout.width_by_header[header] or 18.0)
    if header in JSON_WIDTH_HEADERS or header.startswith("summary.") or header.startswith("calculated."):
        return 52.0
    if header == "address":
        return 52.0
    if header == "source_gcs_uri":
        return 72.0
    if header in {"error", "warning"}:
        return 60.0
    if header in {"source_file_type", "normalized_file_type", "request_file_type", "source_folder", "source_assignee"}:
        return 24.0
    if header in {
        "fixture_status_from_source",
        "failure_tag",
        "error_type",
        "output_generated_at",
        "batch_retry_reason",
        "batch_final_attempt_error_type",
    }:
        return 22.0
    if header in {
        "source_row",
        "batch_chunk_number",
        "batch_result_index",
        "batch_http_status",
        "batch_attempt_count",
    }:
        return 14.0
    if header in {"ok"}:
        return 10.0
    return float(min(max(len(header) + 4, 12), 32))


def _excel_cell_value(value: Any) -> Any:
    if value in ({}, [], ()):
        return None
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return value


def _apply_header_style(cell: Any, layout: TemplateLayout) -> None:
    cell.font = copy(layout.header_font)
    cell.fill = copy(layout.header_fill)
    cell.border = copy(layout.header_border)
    cell.alignment = copy(layout.header_alignment)
    cell.number_format = layout.header_number_format
    cell.protection = copy(layout.header_protection)


def _apply_body_style(cell: Any, layout: TemplateLayout) -> None:
    cell.font = copy(layout.body_font)
    cell.fill = copy(layout.body_fill)
    cell.border = copy(layout.body_border)
    cell.alignment = copy(layout.body_alignment)
    cell.number_format = layout.body_number_format
    cell.protection = copy(layout.body_protection)


def _write_sheet(
    *,
    sheet: Any,
    headers: list[str],
    row_values: list[dict[str, Any]],
    layout: TemplateLayout,
) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_index, header)
        _apply_header_style(cell, layout)
        sheet.column_dimensions[get_column_letter(column_index)].width = _column_width(header, layout)

    for row_index, values in enumerate(row_values, start=2):
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row_index, column_index, _excel_cell_value(values.get(header)))
            _apply_body_style(cell, layout)


def write_workbook(
    *,
    file_type: str,
    rows: list[ExportRow],
    layout: TemplateLayout,
    output_path: Path,
) -> list[str]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_title(file_type)
    sheet.freeze_panes = layout.freeze_panes

    headers = build_main_sheet_header_order(layout, rows)
    analyst_rows: list[dict[str, Any]] = []
    for export_row in rows:
        values: dict[str, Any] = {}
        values.update(export_row.template_values)
        values.update(export_row.extra_values)
        analyst_rows.append(values)
    _write_sheet(sheet=sheet, headers=headers, row_values=analyst_rows, layout=layout)

    meta_sheet = workbook.create_sheet(META_SHEET_TITLE)
    meta_sheet.freeze_panes = layout.freeze_panes or "A2"
    _write_sheet(
        sheet=meta_sheet,
        headers=list(FIXED_METADATA_HEADERS),
        row_values=[export_row.metadata for export_row in rows],
        layout=layout,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return headers


def workbook_filename_for(file_type: str) -> str:
    return f"{_safe_filename(file_type)}__batch_ground_truth.xlsx"
