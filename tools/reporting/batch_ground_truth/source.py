from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tests.endpoints.parse.file_types import request_file_type_for
from tests.endpoints.parse.fixture_json import unsupported_fixture_reason
from tests.fixtures.registry import REGISTRY_PATH as SHARED_REGISTRY_PATH
from tests.fixtures.registry import load_registry
from tools.generate_fixture_registry import (
    COMPOSITE_DELIMITER,
    EXCLUDED_FILE_TYPES,
    HEADER_ROW,
    SHEET_NAME,
    SOURCE_XLSX,
    classify,
    fixture_metadata_overrides_for,
)

from .models import ExcludedSourceRow, SourceFixtureRecord, SourceRegistryParseResult

REPO_ROOT = Path(__file__).resolve().parents[3]
EXPECTED_HEADERS = (
    "Folder",
    "fileType",
    "gsutil Path",
    "fileType Status",
    "Assignee",
    "Status",
)


def _clean_string(value: object) -> str:
    if not isinstance(value, str):
        return ""
    return value.strip()


def _optional_string(value: Any) -> str | None:
    cleaned = _clean_string(value)
    return cleaned or None


def _base_name_from_uri(uri: str) -> str:
    base = uri.rsplit("/", 1)[-1]
    stem = base.rsplit(".", 1)[0] if "." in base else base
    return stem.strip() or "fixture"


def _source_path_from_doc(doc: dict[str, Any]) -> Path | None:
    source = doc.get("source")
    if not isinstance(source, str) or not source.strip():
        return None
    path = Path(source).expanduser()
    if not path.is_absolute():
        path = REPO_ROOT / path
    return path.resolve()


def _status_from_verification(verification_status: str) -> str:
    if verification_status == "confirmed":
        return "✓"
    if verification_status == "unverified":
        return "⚠ Verify"
    return ""


def _split_index(fixture: dict[str, Any]) -> int:
    raw_file_type = fixture.get("source_file_type")
    file_type = fixture.get("file_type")
    if not isinstance(raw_file_type, str) or not isinstance(file_type, str):
        return 0
    parts = [part.strip() for part in raw_file_type.split(COMPOSITE_DELIMITER) if part.strip()]
    try:
        return parts.index(file_type.strip())
    except ValueError:
        return 0


def _source_order_key(fixture: dict[str, Any]) -> tuple[int, int, str, str, str]:
    source_row = fixture.get("source_row")
    if not isinstance(source_row, int) or isinstance(source_row, bool) or source_row <= 0:
        source_row = 1_000_000_000
    return (
        source_row,
        _split_index(fixture),
        str(fixture.get("source_folder") or ""),
        str(fixture.get("file_type") or ""),
        str(fixture.get("name") or ""),
    )


def _registry_record(
    fixture: dict[str, Any],
    *,
    record_id: int,
) -> SourceFixtureRecord:
    gcs_uri = str(fixture["gcs_uri"])
    file_type = str(fixture["file_type"]).strip()
    verification_status = str(fixture["verification_status"]).strip()
    file_type_status = (
        _clean_string(fixture.get("source_file_type_status"))
        or _status_from_verification(verification_status)
    )
    skip_reason = _clean_string(fixture.get("fixture_unsupported_reason")) or unsupported_fixture_reason(gcs_uri)

    return SourceFixtureRecord(
        record_id=record_id,
        source_row=int(fixture["source_row"]),
        source_folder=_optional_string(fixture.get("source_folder")),
        source_file_type=_optional_string(fixture.get("source_file_type")),
        file_type=file_type,
        request_file_type=request_file_type_for(file_type),
        gcs_uri=gcs_uri,
        source_basename=_base_name_from_uri(gcs_uri),
        file_type_status=file_type_status,
        workflow_status=_clean_string(fixture.get("source_workflow_status")),
        assignee=_clean_string(fixture.get("source_assignee")),
        verification_status=verification_status,
        include_in_batch=skip_reason is None,
        skip_reason=skip_reason,
        batch_expected_warning=_optional_string(fixture.get("batch_expected_warning")),
        batch_expected_error_type=_optional_string(fixture.get("batch_expected_error_type")),
        batch_expected_error=_optional_string(fixture.get("batch_expected_error")),
    )


def parse_source_registry(
    fixture_registry: Path | str = SHARED_REGISTRY_PATH,
) -> SourceRegistryParseResult:
    """Load batch planning records from the shared generated YAML registry."""
    registry_path = Path(fixture_registry).expanduser().resolve()
    doc = load_registry(registry_path)
    fixtures: list[SourceFixtureRecord] = []
    excluded_rows: list[ExcludedSourceRow] = []
    record_id = 0

    for fixture in sorted(doc["fixtures"], key=_source_order_key):
        file_type = fixture.get("file_type")
        raw_file_type = (
            _clean_string(fixture.get("source_file_type"))
            or _clean_string(file_type)
            or "<blank>"
        )
        gcs_uri = _optional_string(fixture.get("gcs_uri"))
        source_row = int(fixture["source_row"])

        if not isinstance(file_type, str) or not file_type.strip():
            excluded_rows.append(
                ExcludedSourceRow(
                    source_row=source_row,
                    raw_file_type=raw_file_type,
                    gcs_uri=gcs_uri,
                    reason="missing_file_type",
                )
            )
            continue

        split_file_type = file_type.strip()
        if split_file_type in EXCLUDED_FILE_TYPES or fixture.get("verification_status") == "excluded":
            excluded_rows.append(
                ExcludedSourceRow(
                    source_row=source_row,
                    raw_file_type=split_file_type,
                    gcs_uri=gcs_uri,
                    reason="excluded_file_type",
                )
            )
            continue

        record_id += 1
        fixtures.append(_registry_record(fixture, record_id=record_id))

    return SourceRegistryParseResult(
        source_registry=registry_path,
        source_workbook=_source_path_from_doc(doc),
        schema_version=int(doc["schema_version"]),
        fixtures=tuple(fixtures),
        excluded_rows=tuple(excluded_rows),
    )


def parse_source_workbook_for_comparison(
    source_workbook: Path | str = SOURCE_XLSX,
) -> SourceRegistryParseResult:
    """Legacy Excel parser retained only to compare registry planning parity."""
    workbook_path = Path(source_workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise RuntimeError(f"Source fixture workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Sheet {SHEET_NAME!r} not found in {workbook_path}; got {workbook.sheetnames}")
    sheet = workbook[SHEET_NAME]

    headers = tuple(sheet.cell(HEADER_ROW, column).value for column in range(1, sheet.max_column + 1))
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise RuntimeError(
            f"Unexpected headers in {workbook_path} row {HEADER_ROW}: "
            f"expected {EXPECTED_HEADERS!r}, got {headers[:len(EXPECTED_HEADERS)]!r}"
        )

    fixtures: list[SourceFixtureRecord] = []
    excluded_rows: list[ExcludedSourceRow] = []
    record_id = 0

    for row_idx in range(HEADER_ROW + 1, sheet.max_row + 1):
        folder = _clean_string(sheet.cell(row_idx, 1).value) or None
        raw_file_type = _clean_string(sheet.cell(row_idx, 2).value)
        gcs_uri = _clean_string(sheet.cell(row_idx, 3).value)
        file_type_status = _clean_string(sheet.cell(row_idx, 4).value)
        assignee = _clean_string(sheet.cell(row_idx, 5).value)
        workflow_status = _clean_string(sheet.cell(row_idx, 6).value)

        if raw_file_type:
            split_types = [part.strip() for part in raw_file_type.split(COMPOSITE_DELIMITER) if part.strip()]
        else:
            split_types = []

        if not split_types:
            excluded_rows.append(
                ExcludedSourceRow(
                    source_row=row_idx,
                    raw_file_type=raw_file_type or "<blank>",
                    gcs_uri=gcs_uri or None,
                    reason="missing_file_type",
                )
            )
            continue

        for split_file_type in split_types:
            verification_status, _enabled = classify(split_file_type, file_type_status)
            if split_file_type in EXCLUDED_FILE_TYPES:
                excluded_rows.append(
                    ExcludedSourceRow(
                        source_row=row_idx,
                        raw_file_type=split_file_type,
                        gcs_uri=gcs_uri or None,
                        reason="excluded_file_type",
                    )
                )
                continue

            if not gcs_uri:
                skip_reason = "missing_gcs_uri"
                request_file_type = request_file_type_for(split_file_type)
                source_basename = f"source-row-{row_idx}"
            elif not gcs_uri.startswith("gs://"):
                skip_reason = "invalid_gcs_uri"
                request_file_type = request_file_type_for(split_file_type)
                source_basename = _base_name_from_uri(gcs_uri)
            else:
                skip_reason = unsupported_fixture_reason(gcs_uri)
                request_file_type = request_file_type_for(split_file_type)
                source_basename = _base_name_from_uri(gcs_uri)

            overrides = fixture_metadata_overrides_for(gcs_uri=gcs_uri, file_type=split_file_type)
            record_id += 1
            fixtures.append(
                SourceFixtureRecord(
                    record_id=record_id,
                    source_row=row_idx,
                    source_folder=folder,
                    source_file_type=raw_file_type or None,
                    file_type=split_file_type,
                    request_file_type=request_file_type,
                    gcs_uri=gcs_uri,
                    source_basename=source_basename,
                    file_type_status=file_type_status,
                    workflow_status=workflow_status,
                    assignee=assignee,
                    verification_status=verification_status,
                    include_in_batch=skip_reason is None,
                    skip_reason=skip_reason,
                    batch_expected_warning=overrides.get("batch_expected_warning"),
                    batch_expected_error_type=overrides.get("batch_expected_error_type"),
                    batch_expected_error=overrides.get("batch_expected_error"),
                )
            )

    return SourceRegistryParseResult(
        source_registry=None,
        source_workbook=workbook_path,
        schema_version=0,
        fixtures=tuple(fixtures),
        excluded_rows=tuple(excluded_rows),
    )


def grouped_fixtures_by_file_type(
    parsed: SourceRegistryParseResult,
    *,
    selected_file_types: set[str] | None = None,
) -> dict[str, list[SourceFixtureRecord]]:
    grouped: dict[str, list[SourceFixtureRecord]] = {}
    for fixture in parsed.fixtures:
        if selected_file_types is not None and fixture.file_type not in selected_file_types:
            continue
        grouped.setdefault(fixture.file_type, []).append(fixture)
    return grouped
