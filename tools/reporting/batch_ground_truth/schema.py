from __future__ import annotations

import json
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ExportRow, TemplateLayout
from .triage import MAIN_WORKBOOK_GT_STATUS_HEADERS

FIXED_METADATA_HEADERS = (
    "source_row",
    "source_gcs_uri",
    "source_file_type",
    "normalized_file_type",
    "request_file_type",
    "source_folder",
    "source_assignee",
    "source_workflow_status",
    "fixture_status_from_source",
    "gt_extraction_eligible",
    "gt_extraction_excluded",
    "gt_extraction_skip_reason",
    "gt_extraction_classification",
    "gt_clean_eligible",
    "negative_audit_useful",
    "gt_recovery_action",
    "batch_chunk_number",
    "batch_result_index",
    "batch_http_status",
    "batch_result_correlation_id",
    "batch_elapsed_ms",
    "batch_attempt_count",
    "batch_retry_reason",
    "batch_final_attempt_error_type",
    "ok",
    "failure_tag",
    "error_type",
    "error",
    "warning",
    "raw_result_json",
    "output_generated_at",
)

REQUIRED_MAIN_SHEET_HEADERS = (
    "filename",
    "parse_success",
    "error",
)

HIDDEN_MAIN_SHEET_HEADERS = {
    "summary_json",
    "raw_response_json",
    "transactions_json",
}

SUMMARY_FIELDS_EXCLUDED_FROM_MAIN_SHEET = {
    "document_type",
    "pageNumber",
}

CALCULATED_FIELDS_EXCLUDED_FROM_MAIN_SHEET = {
    "pageNumber",
}

KNOWN_TEMPLATE_HEADERS = {
    "filename",
    "identified_type",
    "confidence",
    "parse_success",
    "quality_score",
    "completeness_score",
    "error",
    "identify_time_ms",
    "parse_time_ms",
    "fraud_score",
    "is_fraudulent",
    "visual_fraud_detected",
    "visual_risk_score",
    "metadata_fraud_score",
    "fraud_indicators_count",
    "high_confidence_fraud_count",
    "fraud_reasons",
    "fraud_report",
    "transactions_count",
    "transactions_json",
    "allowances",
    "basic_pay",
    "bonus",
    "company_name",
    "employee_id",
    "employee_name",
    "employee_tin",
    "gross_pay",
    "hdmf_pagibig_deduction",
    "net_pay",
    "one_time_earnings",
    "other_deduction_amount",
    "other_earnings",
    "overtime_pay",
    "pay_date",
    "pay_period_end_date",
    "pay_period_start_date",
    "philhealth_contribution_deduction",
    "position",
    "sss_contribution_deduction",
    "sss_number",
    "total_deductions",
    "total_government_contributions",
    "total_loan_repayment_amount",
    "withholding_tax_deduction",
    "summary_json",
    "raw_response_json",
}


def _clean_header(value: object) -> str:
    return str(value).strip()


def load_reference_template(reference_workbook: Path | str) -> TemplateLayout:
    workbook_path = Path(reference_workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise RuntimeError(f"Reference workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path)
    if not workbook.sheetnames:
        raise RuntimeError(f"Reference workbook has no sheets: {workbook_path}")
    sheet = workbook[workbook.sheetnames[0]]

    headers = tuple(
        _clean_header(sheet.cell(1, column).value)
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(1, column).value is not None
    )
    if not headers:
        raise RuntimeError(f"Reference workbook has no headers in row 1: {workbook_path}")

    width_by_header: dict[str, float | None] = {}
    for index, header in enumerate(headers, start=1):
        column_letter = sheet.cell(1, index).column_letter
        width_by_header[header] = sheet.column_dimensions[column_letter].width

    return TemplateLayout(
        reference_workbook=workbook_path,
        reference_sheet_name=sheet.title,
        headers=headers,
        freeze_panes=sheet.freeze_panes,
        header_font=copy(sheet["A1"].font),
        header_fill=copy(sheet["A1"].fill),
        header_border=copy(sheet["A1"].border),
        header_alignment=copy(sheet["A1"].alignment),
        header_number_format=sheet["A1"].number_format,
        header_protection=copy(sheet["A1"].protection),
        body_font=copy(sheet["A2"].font),
        body_fill=copy(sheet["A2"].fill),
        body_border=copy(sheet["A2"].border),
        body_alignment=copy(sheet["A2"].alignment),
        body_number_format=sheet["A2"].number_format,
        body_protection=copy(sheet["A2"].protection),
        width_by_header=width_by_header,
    )


def _json_value(value: Any) -> str | None:
    if value in (None, "", [], {}):
        return None
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=False)


def _has_visible_value(value: Any) -> bool:
    return value not in (None, "", [], {})


def _dig(value: Any, *path: str) -> Any:
    current = value
    for key in path:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def _maybe_sequence(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def _clean_extra_header(*, source: str, key: str, used_headers: set[str]) -> str | None:
    cleaned = key.strip()
    if not cleaned:
        return None
    if source == "summary" and cleaned in SUMMARY_FIELDS_EXCLUDED_FROM_MAIN_SHEET:
        return None
    if source == "calculated" and cleaned in CALCULATED_FIELDS_EXCLUDED_FROM_MAIN_SHEET:
        return None

    candidate = cleaned.replace(".", "_")
    if source == "calculated" and not candidate.startswith("calculated_"):
        candidate = f"calculated_{candidate}"
    if candidate in used_headers:
        candidate = f"{source}_{candidate}"
    return candidate


def build_success_template_values(
    *,
    source_basename: str,
    request_file_type: str,
    result: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any]]:
    data = result.get("data")
    if not isinstance(data, dict):
        raise ValueError("successful batch result is missing a dict `data` payload")

    summary_rows = _maybe_sequence(data.get("summaryResult"))
    if not summary_rows or not isinstance(summary_rows[0], dict):
        raise ValueError("successful batch result is missing a usable `summaryResult[0]` payload")
    summary_row = dict(summary_rows[0])

    calculated_rows = _maybe_sequence(data.get("calculatedFields"))
    first_calculated = calculated_rows[0] if calculated_rows and isinstance(calculated_rows[0], dict) else {}
    fraud_report = _maybe_sequence(data.get("fraudReport"))
    fraud_findings = _maybe_sequence(data.get("fraudCheckFindings"))
    transactions = _maybe_sequence(data.get("transactionsOCR"))
    timings = _dig(data, "timings") if isinstance(_dig(data, "timings"), dict) else {}
    parse_time_ms = None
    if isinstance(timings, dict):
        llm_timing = timings.get("llm_parsing_ms")
        if isinstance(llm_timing, dict):
            parse_time_ms = llm_timing.get("total_ms")
        elif isinstance(llm_timing, (int, float)):
            parse_time_ms = llm_timing

    template_values: dict[str, Any] = {
        "filename": source_basename,
        "identified_type": data.get("fileType") or request_file_type,
        "confidence": data.get("confidence"),
        "parse_success": True,
        "quality_score": data.get("qualityScore"),
        "completeness_score": data.get("completenessScore"),
        "error": None,
        "identify_time_ms": None,
        "parse_time_ms": parse_time_ms,
        "fraud_score": data.get("fraudScore"),
        "is_fraudulent": _dig(data, "mathematicalFraudReport", "is_fraudulent"),
        "visual_fraud_detected": _dig(data, "mathematicalFraudReport", "visual_fraud_detected"),
        "visual_risk_score": _dig(data, "mathematicalFraudReport", "visual_fraud_score"),
        "metadata_fraud_score": _dig(data, "metadataFraudReport", "fraud_score"),
        "fraud_indicators_count": _dig(data, "mathematicalFraudReport", "total_indicators"),
        "high_confidence_fraud_count": _dig(data, "mathematicalFraudReport", "high_confidence_count"),
        "fraud_reasons": _json_value(
            [item.get("description") for item in fraud_findings if isinstance(item, dict) and item.get("description")]
        ),
        "fraud_report": _json_value(fraud_report),
        "transactions_count": len(transactions),
        "transactions_json": _json_value(transactions),
        "summary_json": _json_value(summary_rows),
        "raw_response_json": _json_value(data),
    }

    for key, value in summary_row.items():
        template_values.setdefault(key, value)

    extra_values: dict[str, Any] = {}
    used_headers = set(KNOWN_TEMPLATE_HEADERS)
    for key in sorted(summary_row):
        if key not in KNOWN_TEMPLATE_HEADERS:
            header = _clean_extra_header(source="summary", key=key, used_headers=used_headers | set(extra_values))
            if header is not None:
                extra_values[header] = summary_row[key]
    for key in sorted(first_calculated):
        header = _clean_extra_header(source="calculated", key=key, used_headers=used_headers | set(extra_values))
        if header is not None:
            extra_values[header] = first_calculated[key]

    return template_values, extra_values


def build_failure_template_values(*, source_basename: str, error: str | None) -> dict[str, Any]:
    return {
        "filename": source_basename,
        "parse_success": False,
        "error": error,
    }


def build_main_sheet_header_order(
    layout: TemplateLayout,
    rows: list[ExportRow],
    *,
    include_gt_status_columns: bool = False,
) -> list[str]:
    template_headers = [
        header
        for header in layout.headers
        if header not in HIDDEN_MAIN_SHEET_HEADERS
        and (
            header in REQUIRED_MAIN_SHEET_HEADERS
            or any(_has_visible_value(row.template_values.get(header)) for row in rows)
        )
    ]
    gt_status_headers = [
        header
        for header in MAIN_WORKBOOK_GT_STATUS_HEADERS
        if include_gt_status_columns and header not in template_headers
    ]
    dynamic_headers = sorted(
        {
            header
            for row in rows
            for header, value in row.extra_values.items()
            if _has_visible_value(value) and header not in gt_status_headers
        }
    )
    return [*template_headers, *gt_status_headers, *dynamic_headers]
