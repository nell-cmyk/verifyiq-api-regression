from __future__ import annotations

import csv
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from copy import copy
import json
from pathlib import Path
from types import SimpleNamespace
import threading
import time

import httpx
import pytest

from openpyxl import Workbook, load_workbook

import tools.reporting.batch_ground_truth.workflow as batch_ground_truth_workflow
import tools.reporting.export_batch_ground_truth as export_batch_ground_truth
from tests.endpoints.batch.artifacts import (
    BATCH_RESPONSE_ARTIFACT_DIR_ENV_VAR,
    BATCH_RESPONSE_ARTIFACT_RUN_DIR_ENV_VAR,
)
from tools.generate_fixture_registry import build_registry_document, write_registry_document
from tools.reporting.batch_ground_truth.excel import clean_workbook_filename_for, write_workbook
from tools.reporting.batch_ground_truth.models import ExportRow, FileTypePlan, SourceFixtureRecord
from tools.reporting.batch_ground_truth.schema import (
    FIXED_METADATA_HEADERS,
    build_success_template_values,
    load_reference_template,
)
from tools.reporting.batch_ground_truth.source import (
    parse_source_registry,
    parse_source_workbook_for_comparison,
)
from tools.reporting.batch_ground_truth.workflow import plan_file_types
from tools.reporting.batch_ground_truth.triage import (
    build_recovery_triage_row,
    classify_export_row,
    is_clean_candidate,
)


def _write_source_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "All"
    sheet["A1"] = "QA Fixture Registry — POST /v1/documents/parse & /batch"
    sheet["A2"] = "Source: gs://verifyiq-internal-testing/QA/GroundTruth/"
    sheet["A4"] = "Folder"
    sheet["B4"] = "fileType"
    sheet["C4"] = "gsutil Path"
    sheet["D4"] = "fileType Status"
    sheet["E4"] = "Assignee"
    sheet["F4"] = "Status"

    rows = [
        (
            "Tax",
            "BIRForm2303 || BIRExemptionCertificate",
            "gs://verifyiq-internal-testing/QA/GroundTruth/Tax/example.pdf",
            "⚠ Verify",
            "Thor",
            "Pending",
        ),
        (
            "BankStatement",
            "BankStatement",
            "gs://verifyiq-internal-testing/QA/GroundTruth/BankStatement/example.png.xlsx",
            "✓",
            "Jane",
            "Pending",
        ),
        (
            "BankStatement",
            "BankStatement",
            (
                "gs://verifyiq-internal-testing/QA/GroundTruth/BankStatement/"
                "MJRL_MV Dela Cruz_Bank Statement (1).pdf"
            ),
            "✓",
            "Jane",
            "Pending",
        ),
        (
            "Missing",
            "No fileType",
            "gs://verifyiq-internal-testing/QA/GroundTruth/Missing/example.pdf",
            "✓",
            "Alex",
            "Pending",
        ),
        (
            "Fraud",
            "Fraud - Skipped",
            "gs://verifyiq-internal-testing/QA/GroundTruth/Fraud/example.pdf",
            "✓",
            "Alex",
            "Skipped",
        ),
    ]
    for row_index, row in enumerate(rows, start=5):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row_index, column_index, value)

    workbook.save(path)
    return path


def _write_fixture_registry_from_source_workbook(source_path: Path, registry_path: Path) -> Path:
    doc = build_registry_document(source_xlsx=source_path, supplemental_yaml=None)
    write_registry_document(doc, output_paths=(registry_path,))
    return registry_path


def _write_fixture_registry(
    registry_path: Path,
    *,
    file_type_counts: dict[str, int],
) -> Path:
    fixtures: list[dict[str, object]] = []
    source_row = 5
    for file_type, count in file_type_counts.items():
        for index in range(1, count + 1):
            fixtures.append(
                {
                    "name": f"{file_type.lower()}-{index}",
                    "file_type": file_type,
                    "gcs_uri": f"gs://bucket/{file_type.lower()}-{index}.pdf",
                    "source_folder": file_type,
                    "source_file_type": file_type,
                    "source_file_type_status": "✓",
                    "source_assignee": "Thor",
                    "source_workflow_status": "Pending",
                    "source_row": source_row,
                    "verification_status": "confirmed",
                    "enabled": True,
                }
            )
            source_row += 1

    write_registry_document(
        {
            "schema_version": 2,
            "total": len(fixtures),
            "composite_rows_split": 0,
            "counts": {"confirmed": len(fixtures)},
            "fixtures": fixtures,
        },
        output_paths=(registry_path,),
    )
    return registry_path


def _write_mixed_fixture_registry(registry_path: Path) -> Path:
    fixtures = [
        {
            "name": "clean",
            "file_type": "Payslip",
            "gcs_uri": "gs://bucket/clean.pdf",
            "source_folder": "Payslip",
            "source_file_type": "Payslip",
            "source_file_type_status": "✓",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "source_row": 10,
            "verification_status": "confirmed",
            "enabled": True,
        },
        {
            "name": "quality-gated",
            "file_type": "Payslip",
            "gcs_uri": "gs://bucket/quality-gated.pdf",
            "source_folder": "Payslip",
            "source_file_type": "Payslip",
            "source_file_type_status": "✓",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "source_row": 11,
            "verification_status": "confirmed",
            "enabled": True,
        },
        {
            "name": "size-guard",
            "file_type": "Payslip",
            "gcs_uri": "gs://bucket/size-guard.pdf",
            "source_folder": "Payslip",
            "source_file_type": "Payslip",
            "source_file_type_status": "✓",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "source_row": 12,
            "verification_status": "confirmed",
            "enabled": True,
            "batch_expected_error_type": "DocumentSizeGuardError",
            "batch_expected_error": "document exceeded size guard",
        },
        {
            "name": "unsupported",
            "file_type": "Payslip",
            "gcs_uri": "gs://bucket/unsupported.xlsx",
            "source_folder": "Payslip",
            "source_file_type": "Payslip",
            "source_file_type_status": "✓",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "source_row": 13,
            "verification_status": "confirmed",
            "enabled": True,
        },
    ]
    write_registry_document(
        {
            "schema_version": 2,
            "total": len(fixtures),
            "composite_rows_split": 0,
            "counts": {"confirmed": len(fixtures)},
            "fixtures": fixtures,
        },
        output_paths=(registry_path,),
    )
    return registry_path


def _write_gt_exclusion_registry(
    registry_path: Path,
    *,
    excluded_fixture: dict[str, object],
) -> Path:
    clean_fixture = {
        "name": "clean",
        "file_type": "Payslip",
        "gcs_uri": "gs://bucket/clean.pdf",
        "source_folder": "Payslip",
        "source_file_type": "Payslip",
        "source_file_type_status": "✓",
        "source_assignee": "Thor",
        "source_workflow_status": "Pending",
        "source_row": 10,
        "verification_status": "confirmed",
        "enabled": True,
    }
    write_registry_document(
        {
            "schema_version": 2,
            "total": 2,
            "composite_rows_split": 0,
            "counts": {"confirmed": 2},
            "fixtures": [clean_fixture, excluded_fixture],
        },
        output_paths=(registry_path,),
    )
    return registry_path


def _write_reference_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Payslip"
    headers = [
        "filename",
        "identified_type",
        "parse_success",
        "error",
        "summary_json",
        "raw_response_json",
    ]
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(1, index, header)
        font = copy(cell.font)
        font.bold = True
        cell.font = font
        sheet.column_dimensions[cell.column_letter].width = 24
    for index in range(1, len(headers) + 1):
        sheet.cell(2, index, None)
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path


def _make_fixture(
    record_id: int,
    *,
    file_type: str = "Payslip",
    include_in_batch: bool = True,
    skip_reason: str | None = None,
    batch_expected_error_type: str | None = None,
    batch_expected_warning: str | None = None,
    gt_extraction_eligible: bool = True,
    gt_extraction_skip_reason: str | None = None,
    gt_extraction_classification: str | None = None,
    gt_recovery_action: str | None = None,
    gt_clean_eligible: bool = True,
    negative_audit_useful: bool = False,
) -> SourceFixtureRecord:
    return SourceFixtureRecord(
        record_id=record_id,
        source_row=100 + record_id,
        source_folder=file_type,
        source_file_type=file_type,
        file_type=file_type,
        request_file_type=file_type,
        gcs_uri=f"gs://bucket/fixture-{record_id}.pdf",
        source_basename=f"fixture-{record_id}.pdf",
        file_type_status="✓",
        workflow_status="Pending",
        assignee="Thor",
        verification_status="verified",
        include_in_batch=include_in_batch,
        skip_reason=skip_reason,
        batch_expected_warning=batch_expected_warning,
        batch_expected_error_type=batch_expected_error_type,
        batch_expected_error="expected warning result" if batch_expected_error_type else None,
        gt_extraction_eligible=gt_extraction_eligible,
        gt_extraction_skip_reason=gt_extraction_skip_reason,
        gt_extraction_classification=gt_extraction_classification,
        gt_recovery_action=gt_recovery_action,
        gt_clean_eligible=gt_clean_eligible,
        negative_audit_useful=negative_audit_useful,
    )


def _success_result(record_id: int, index: int) -> dict[str, object]:
    return {
        "index": index,
        "ok": True,
        "correlation_id": f"corr-{record_id}",
        "elapsed_ms": 10.0 + index,
        "data": {
            "fileType": "Payslip",
            "summaryResult": [
                {
                    "document_type": "Payslip",
                    "employee_name": f"Employee {record_id}",
                }
            ],
            "calculatedFields": [
                {
                    "pageNumber": 1,
                    "calculated.match_score": 90 + index,
                }
            ],
        },
    }


def _error_result(
    index: int,
    *,
    error_type: str,
    error: str,
    warning: str | None = None,
) -> dict[str, object]:
    result: dict[str, object] = {
        "index": index,
        "ok": False,
        "error_type": error_type,
        "error": error,
    }
    if warning is not None:
        result["warning"] = warning
    return result


def _unusable_success_result(index: int) -> dict[str, object]:
    return {
        "index": index,
        "ok": True,
        "correlation_id": f"corr-unusable-{index}",
        "elapsed_ms": 33.3,
        "data": {
            "fileType": "Payslip",
        },
    }


def _quality_gated_no_payload_result(index: int) -> dict[str, object]:
    return {
        "index": index,
        "ok": True,
        "correlation_id": f"corr-quality-gated-{index}",
        "elapsed_ms": 44.4,
        "data": {
            "fileType": "Payslip",
            "summaryOCR": [],
            "summaryResult": [],
            "calculatedFields": [],
            "transactionsOCR": [],
            "documentQuality": "Failed: failed document_classification",
            "qualityScore": 0.0,
            "qualityCheck": {
                "issueDescription": "Failed: failed document_classification",
                "qualityCheckFindings": [
                    {
                        "type": "document_classification",
                        "status": "failed",
                        "description": "Image does not appear to contain a valid document.",
                    }
                ],
            },
            "extractionStatus": "not_attempted",
        },
    }


def _unknown_no_payload_result(index: int) -> dict[str, object]:
    return {
        "index": index,
        "ok": True,
        "correlation_id": f"corr-empty-{index}",
        "elapsed_ms": 45.5,
        "data": {
            "fileType": "Payslip",
            "summaryOCR": [],
            "summaryResult": [],
            "calculatedFields": [],
            "transactionsOCR": [],
            "extractionStatus": "not_attempted",
        },
    }


def _null_field_success_result(record_id: int, index: int) -> dict[str, object]:
    result = _success_result(record_id, index)
    data = result["data"]
    assert isinstance(data, dict)
    data["summaryResult"] = [
        {
            "document_type": "Payslip",
            "employee_name": None,
            "net_pay": None,
        }
    ]
    return result


def _chunk_key(fixtures: list[SourceFixtureRecord]) -> tuple[str, ...]:
    return tuple(fixture.gcs_uri for fixture in fixtures)


def _registry_chunk_key(file_type: str, indexes: range) -> tuple[str, ...]:
    return tuple(f"gs://bucket/{file_type.lower()}-{index}.pdf" for index in indexes)


def _success_body_for_key(key: tuple[str, ...]) -> dict[str, object]:
    return {
        "results": [
            _success_result(record_id=index + 1, index=index)
            for index, _gcs_uri in enumerate(key)
        ]
    }


def _configure_batch_artifact_dir(monkeypatch, tmp_path: Path) -> Path:
    artifact_root = tmp_path / "batch-artifacts"
    monkeypatch.setenv(BATCH_RESPONSE_ARTIFACT_DIR_ENV_VAR, str(artifact_root))
    monkeypatch.setenv(BATCH_RESPONSE_ARTIFACT_RUN_DIR_ENV_VAR, "batch-test-run")
    return artifact_root / "batch-test-run"


def _install_fake_make_client(monkeypatch, plans: dict[tuple[str, ...], object]) -> dict[str, object]:
    state: dict[str, object] = {
        "client_count": 0,
        "force_refresh_count": 0,
        "closed_count": 0,
        "attempts": {},
        "active": 0,
        "max_active": 0,
        "completion_order": [],
    }
    lock = threading.Lock()

    class FakeBatchClient:
        def __enter__(self) -> FakeBatchClient:
            return self

        def __exit__(self, exc_type, exc, tb) -> bool:
            return False

        def close(self) -> None:
            state["closed_count"] = int(state["closed_count"]) + 1

        def post(self, url: str, *, json: dict[str, object] | None = None, timeout: float | None = None) -> httpx.Response:
            assert url == batch_ground_truth_workflow.BATCH_ENDPOINT
            assert timeout == batch_ground_truth_workflow.BATCH_TIMEOUT_SECS
            assert json is not None

            key = tuple(str(item["file"]) for item in json["items"])
            request = httpx.Request("POST", f"https://example.test{url}", json=json)

            with lock:
                attempts = state["attempts"]
                assert isinstance(attempts, dict)
                attempt_index = int(attempts.get(key, 0))
                attempts[key] = attempt_index + 1

            raw_plan = plans[key]
            if isinstance(raw_plan, list):
                plan = raw_plan[min(attempt_index, len(raw_plan) - 1)]
            else:
                plan = raw_plan
            assert isinstance(plan, dict)

            with lock:
                state["active"] = int(state["active"]) + 1
                state["max_active"] = max(int(state["max_active"]), int(state["active"]))

            try:
                barrier = plan.get("barrier")
                if barrier is not None:
                    assert isinstance(barrier, threading.Barrier)
                    barrier.wait(timeout=5.0)

                delay = float(plan.get("delay", 0.0))
                if delay:
                    time.sleep(delay)

                name = str(plan.get("name", key[0]))
                cast_completion_order = state["completion_order"]
                assert isinstance(cast_completion_order, list)
                cast_completion_order.append(name)

                kind = str(plan["kind"])
                if kind == "timeout":
                    raise httpx.ReadTimeout("timed out", request=request)
                if kind == "remote_protocol":
                    raise httpx.RemoteProtocolError("server disconnected", request=request)
                if kind == "request_error":
                    raise httpx.ConnectError("connection failed", request=request)
                if kind == "invalid_json":
                    return httpx.Response(
                        int(plan.get("status", 200)),
                        text=str(plan.get("text", "not-json")),
                        headers=dict(plan.get("headers", {})),
                        request=request,
                    )
                return httpx.Response(
                    int(plan.get("status", 200)),
                    json=plan["body"],
                    headers=dict(plan.get("headers", {})),
                    request=request,
                )
            finally:
                with lock:
                    state["active"] = int(state["active"]) - 1

    def fake_make_client(
        timeout: float = 60.0,
        force_refresh_iap_token: bool = False,
    ) -> FakeBatchClient:
        assert timeout == batch_ground_truth_workflow.BATCH_TIMEOUT_SECS
        state["client_count"] = int(state["client_count"]) + 1
        if force_refresh_iap_token:
            state["force_refresh_count"] = int(state["force_refresh_count"]) + 1
        return FakeBatchClient()

    monkeypatch.setattr(batch_ground_truth_workflow, "make_client", fake_make_client)
    return state


def _row_signatures(rows: list[ExportRow]) -> list[tuple[dict[str, object], dict[str, object], dict[str, object]]]:
    return [
        (
            dict(row.metadata),
            dict(row.template_values),
            dict(row.extra_values),
        )
        for row in rows
    ]


def test_parse_source_workbook_comparison_helper_splits_and_keeps_verify_rows(tmp_path):
    source_path = _write_source_workbook(tmp_path / "qa_fixture_registry.xlsx")

    parsed = parse_source_workbook_for_comparison(source_path)

    assert parsed.source_workbook == source_path.resolve()
    assert len(parsed.fixtures) == 4
    split_types = [fixture.file_type for fixture in parsed.fixtures]
    assert split_types == [
        "BIRForm2303",
        "BIRExemptionCertificate",
        "BankStatement",
        "BankStatement",
    ]

    bir_fixture = parsed.fixtures[0]
    assert bir_fixture.include_in_batch is True
    assert bir_fixture.file_type_status == "⚠ Verify"
    assert bir_fixture.verification_status == "unverified"
    assert bir_fixture.source_file_type == "BIRForm2303 || BIRExemptionCertificate"

    bank_fixture = parsed.fixtures[2]
    assert bank_fixture.include_in_batch is False
    assert bank_fixture.skip_reason == "unsupported file extension '.xlsx' (supported: PDF, PNG, JPG, JPEG, TIFF, TIF, HEIC, HEIF)"
    assert bank_fixture.gt_extraction_eligible is False
    assert bank_fixture.gt_extraction_skip_reason == "unsupported_fixture"
    assert bank_fixture.gt_extraction_classification == "unsupported_artifact"
    assert bank_fixture.gt_recovery_action == "replace_fixture"

    warning_fixture = parsed.fixtures[3]
    assert warning_fixture.include_in_batch is False
    assert warning_fixture.batch_expected_error_type == "DocumentSizeGuardError"
    assert "Page count (456)" in str(warning_fixture.batch_expected_error)
    assert "GT exclusion" in str(warning_fixture.batch_expected_warning)
    assert warning_fixture.gt_extraction_eligible is False
    assert warning_fixture.gt_extraction_skip_reason == "document_size_guard"
    assert warning_fixture.gt_extraction_classification == "fixture_too_large"
    assert warning_fixture.gt_recovery_action == "reduce_fixture"

    assert len(parsed.excluded_rows) == 2
    assert parsed.excluded_rows[0].raw_file_type == "No fileType"
    assert parsed.excluded_rows[0].reason == "excluded_file_type"
    assert parsed.excluded_rows[1].raw_file_type == "Fraud - Skipped"
    assert parsed.excluded_rows[1].reason == "excluded_file_type"


def test_registry_batch_planning_matches_legacy_workbook_parser(tmp_path):
    source_path = _write_source_workbook(tmp_path / "qa_fixture_registry.xlsx")
    registry_path = _write_fixture_registry_from_source_workbook(
        source_path,
        tmp_path / "fixture_registry.yaml",
    )

    legacy = parse_source_workbook_for_comparison(source_path)
    registry = parse_source_registry(registry_path)

    def fixture_signature(fixture: SourceFixtureRecord) -> tuple[object, ...]:
        return (
            fixture.source_row,
            fixture.source_folder,
            fixture.source_file_type,
            fixture.file_type,
            fixture.request_file_type,
            fixture.gcs_uri,
            fixture.source_basename,
            fixture.file_type_status,
            fixture.workflow_status,
            fixture.assignee,
            fixture.verification_status,
            fixture.include_in_batch,
            fixture.skip_reason,
            fixture.batch_expected_warning,
            fixture.batch_expected_error_type,
            fixture.batch_expected_error,
            fixture.gt_extraction_eligible,
            fixture.gt_extraction_skip_reason,
            fixture.gt_extraction_classification,
            fixture.gt_recovery_action,
            fixture.gt_clean_eligible,
            fixture.negative_audit_useful,
        )

    assert [fixture_signature(fixture) for fixture in registry.fixtures] == [
        fixture_signature(fixture) for fixture in legacy.fixtures
    ]
    assert registry.excluded_rows == legacy.excluded_rows


def test_plan_file_types_reports_executable_and_skipped_counts(tmp_path):
    source_path = _write_source_workbook(tmp_path / "qa_fixture_registry.xlsx")
    registry_path = _write_fixture_registry_from_source_workbook(
        source_path,
        tmp_path / "fixture_registry.yaml",
    )

    parsed, grouped, plans = plan_file_types(fixture_registry=registry_path)

    assert len(parsed.fixtures) == 4
    assert sorted(grouped) == ["BIRExemptionCertificate", "BIRForm2303", "BankStatement"]

    plan_by_type = {plan.file_type: plan for plan in plans}
    assert plan_by_type["BIRForm2303"].executable_rows == 1
    assert plan_by_type["BankStatement"].total_rows == 2
    assert plan_by_type["BankStatement"].executable_rows == 0
    assert plan_by_type["BankStatement"].skipped_rows == 2

    _selected_parsed, selected_grouped, selected_plans = plan_file_types(
        fixture_registry=registry_path,
        selected_file_types={"BankStatement"},
    )
    assert sorted(selected_grouped) == ["BankStatement"]
    assert [plan.file_type for plan in selected_plans] == ["BankStatement"]


def test_build_success_template_values_maps_common_and_extra_fields():
    result = {
        "index": 0,
        "ok": True,
        "data": {
            "fileType": "TINID",
            "qualityScore": 91.2,
            "completenessScore": 97,
            "fraudScore": 5.5,
            "timings": {"llm_parsing_ms": {"total_ms": 1234.5}},
            "fraudReport": ["reason-a"],
            "fraudCheckFindings": [
                {"description": "reason-a"},
                {"description": "reason-b"},
            ],
            "metadataFraudReport": {"fraud_score": 2},
            "mathematicalFraudReport": {
                "is_fraudulent": False,
                "visual_fraud_detected": False,
                "visual_fraud_score": 0,
                "total_indicators": 1,
                "high_confidence_count": 0,
            },
            "transactionsOCR": [{"id": 1}],
            "summaryResult": [
                {
                    "document_type": "TINID",
                    "tin": "123-456-789",
                    "pageNumber": 1,
                    "custom_extra_field": "extra",
                }
            ],
            "calculatedFields": [
                {
                    "pageNumber": 1,
                    "calculated.match_score": 88,
                }
            ],
        },
    }

    template_values, extra_values = build_success_template_values(
        source_basename="fixture-name",
        request_file_type="TINID",
        result=result,
    )

    assert template_values["filename"] == "fixture-name"
    assert template_values["identified_type"] == "TINID"
    assert template_values["parse_success"] is True
    assert template_values["quality_score"] == 91.2
    assert template_values["parse_time_ms"] == 1234.5
    assert template_values["transactions_count"] == 1
    assert template_values["fraud_reasons"] == '["reason-a","reason-b"]'
    assert template_values["raw_response_json"].startswith('{"fileType":"TINID"')
    assert extra_values["custom_extra_field"] == "extra"
    assert extra_values["calculated_match_score"] == 88
    assert "document_type" not in extra_values
    assert "pageNumber" not in extra_values


def test_write_workbook_keeps_main_sheet_analyst_facing_and_writes_meta_sheet(tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    layout = load_reference_template(reference_path)

    success_row = ExportRow(
        metadata={
            "source_row": 10,
            "source_gcs_uri": "gs://bucket/success.pdf",
            "source_file_type": "Payslip",
            "normalized_file_type": "Payslip",
            "request_file_type": "Payslip",
            "source_folder": "Payslip",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "fixture_status_from_source": "✓",
            "batch_chunk_number": 1,
            "batch_result_index": 0,
            "batch_http_status": 200,
            "batch_result_correlation_id": "corr-1",
            "batch_elapsed_ms": 12.5,
            "batch_attempt_count": 1,
            "batch_retry_reason": None,
            "batch_final_attempt_error_type": None,
            "ok": True,
            "failure_tag": None,
            "error_type": None,
            "error": None,
            "warning": None,
            "raw_result_json": '{"ok":true}',
            "output_generated_at": "2026-04-23T00:00:00+00:00",
        },
        template_values={
            "filename": "success.pdf",
            "identified_type": "Payslip",
            "parse_success": True,
            "error": None,
            "summary_json": '[{"document_type":"Payslip"}]',
            "raw_response_json": '{"fileType":"Payslip"}',
        },
        extra_values={"custom_field": "custom"},
    )
    failure_row = ExportRow(
        metadata={
            "source_row": 11,
            "source_gcs_uri": "gs://bucket/failure.pdf",
            "source_file_type": "Payslip",
            "normalized_file_type": "Payslip",
            "request_file_type": "Payslip",
            "source_folder": "Payslip",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "fixture_status_from_source": "⚠ Verify",
            "batch_chunk_number": None,
            "batch_result_index": None,
            "batch_http_status": None,
            "batch_result_correlation_id": None,
            "batch_elapsed_ms": None,
            "batch_attempt_count": None,
            "batch_retry_reason": None,
            "batch_final_attempt_error_type": None,
            "ok": False,
            "failure_tag": "unsupported_fixture",
            "error_type": None,
            "error": "unsupported file extension '.xlsx'",
            "warning": None,
            "raw_result_json": None,
            "output_generated_at": "2026-04-23T00:00:00+00:00",
        },
        template_values={
            "filename": "failure.pdf",
            "parse_success": False,
            "error": "unsupported file extension '.xlsx'",
        },
    )

    output_path = tmp_path / "Payslip__batch_ground_truth.xlsx"
    headers = write_workbook(
        file_type="Payslip",
        rows=[success_row, failure_row],
        layout=layout,
        output_path=output_path,
    )

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Payslip"]
    meta_sheet = workbook["_meta"]
    main_headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    meta_headers = [meta_sheet.cell(1, column).value for column in range(1, meta_sheet.max_column + 1)]

    assert workbook.sheetnames == ["Payslip", "_meta"]
    assert sheet.freeze_panes == "A2"
    assert headers == ["filename", "identified_type", "parse_success", "error", "custom_field"]
    assert main_headers == headers
    assert "source_gcs_uri" not in headers
    assert "raw_result_json" not in headers
    assert Counter(headers)["error"] == 1

    filename_column = headers.index("filename") + 1
    parse_success_column = headers.index("parse_success") + 1
    error_column = headers.index("error") + 1
    assert sheet.cell(2, filename_column).value == "success.pdf"
    assert sheet.cell(3, filename_column).value == "failure.pdf"
    assert sheet.cell(3, parse_success_column).value is False
    assert sheet.cell(3, error_column).value == "unsupported file extension '.xlsx'"
    assert sheet.freeze_panes == "A2"

    assert meta_sheet.freeze_panes == "A2"
    assert meta_headers == list(FIXED_METADATA_HEADERS)
    source_uri_column = meta_headers.index("source_gcs_uri") + 1
    raw_result_column = meta_headers.index("raw_result_json") + 1
    assert meta_sheet.cell(2, source_uri_column).value == "gs://bucket/success.pdf"
    assert meta_sheet.cell(3, source_uri_column).value == "gs://bucket/failure.pdf"
    assert meta_sheet.cell(2, raw_result_column).value == '{"ok":true}'
    assert meta_sheet.cell(3, raw_result_column).value is None


def test_write_workbook_serializes_list_valued_extra_fields(tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    layout = load_reference_template(reference_path)

    row = ExportRow(
        metadata={
            "source_row": 10,
            "source_gcs_uri": "gs://bucket/list-value.pdf",
            "source_file_type": "Payslip",
            "normalized_file_type": "Payslip",
            "request_file_type": "Payslip",
            "source_folder": "Payslip",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "fixture_status_from_source": "✓",
            "batch_chunk_number": 1,
            "batch_result_index": 0,
            "batch_http_status": 200,
            "batch_result_correlation_id": "corr-1",
            "batch_elapsed_ms": 12.5,
            "batch_attempt_count": 1,
            "batch_retry_reason": None,
            "batch_final_attempt_error_type": None,
            "ok": True,
            "failure_tag": None,
            "error_type": None,
            "error": None,
            "warning": None,
            "raw_result_json": '{"ok":true}',
            "output_generated_at": "2026-04-23T00:00:00+00:00",
        },
        template_values={
            "filename": "list-value.pdf",
            "identified_type": "Payslip",
            "parse_success": True,
            "error": None,
        },
        extra_values={"income_types": ["REMITTANCE"]},
    )

    output_path = tmp_path / "Payslip__batch_ground_truth.xlsx"
    headers = write_workbook(
        file_type="Payslip",
        rows=[row],
        layout=layout,
        output_path=output_path,
    )

    workbook = load_workbook(output_path, data_only=False)
    sheet = workbook["Payslip"]
    income_types_column = headers.index("income_types") + 1

    assert sheet.cell(2, income_types_column).value == '["REMITTANCE"]'


def test_execute_file_type_preserves_default_sequential_behavior(monkeypatch):
    fixtures = [_make_fixture(record_id) for record_id in range(1, 9)]
    chunk_one = fixtures[:4]
    chunk_two = fixtures[4:]
    plans = {
        _chunk_key(chunk_one): {
            "kind": "json",
            "body": {"results": [_success_result(record_id, index) for index, record_id in enumerate(range(1, 5))]},
            "delay": 0.02,
            "name": "chunk-1",
        },
        _chunk_key(chunk_two): {
            "kind": "json",
            "body": {"results": [_success_result(record_id, index) for index, record_id in enumerate(range(5, 9))]},
            "name": "chunk-2",
        },
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    assert chunk_count == 2
    assert state["client_count"] == 1
    assert state["max_active"] == 1
    assert state["completion_order"] == ["chunk-1", "chunk-2"]
    assert [row.metadata["source_row"] for row in rows] == [fixture.source_row for fixture in fixtures]
    assert [row.metadata["batch_chunk_number"] for row in rows] == [1, 1, 1, 1, 2, 2, 2, 2]


def test_execute_file_type_runs_chunks_concurrently_and_keeps_source_order(monkeypatch):
    fixtures = [
        _make_fixture(record_id) for record_id in range(1, 5)
    ] + [
        _make_fixture(
            5,
            include_in_batch=False,
            skip_reason=(
                "unsupported file extension '.xlsx' "
                "(supported: PDF, PNG, JPG, JPEG, TIFF, TIF, HEIC, HEIF)"
            ),
        )
    ] + [
        _make_fixture(record_id) for record_id in range(6, 10)
    ]
    chunk_one = fixtures[:4]
    chunk_two = fixtures[5:]
    barrier = threading.Barrier(2)
    plans = {
        _chunk_key(chunk_one): {
            "kind": "json",
            "body": {"results": [_success_result(record_id, index) for index, record_id in enumerate(range(1, 5))]},
            "delay": 0.03,
            "barrier": barrier,
            "name": "chunk-1",
        },
        _chunk_key(chunk_two): {
            "kind": "json",
            "body": {"results": [_success_result(record_id, index) for index, record_id in enumerate(range(6, 10))]},
            "barrier": barrier,
            "name": "chunk-2",
        },
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=2,
    )

    assert chunk_count == 2
    assert state["client_count"] == 2
    assert state["max_active"] == 2
    assert state["completion_order"] == ["chunk-2", "chunk-1"]
    assert [row.metadata["source_row"] for row in rows] == [fixture.source_row for fixture in fixtures]
    assert [row.metadata["batch_chunk_number"] for row in rows] == [1, 1, 1, 1, None, 2, 2, 2, 2]
    assert rows[4].metadata["failure_tag"] == "unsupported_fixture"
    assert rows[4].metadata["ok"] is False


def test_execute_file_type_keeps_classification_identical_in_sequential_and_concurrent_modes(monkeypatch):
    fixtures = [_make_fixture(record_id) for record_id in range(1, 26)]
    fixtures[3] = _make_fixture(
        4,
        batch_expected_error_type="DocumentSizeGuardError",
        batch_expected_warning="document exceeded size guard",
    )
    fixtures.append(
        _make_fixture(
            26,
            include_in_batch=False,
            skip_reason=(
                "unsupported file extension '.xlsx' "
                "(supported: PDF, PNG, JPG, JPEG, TIFF, TIF, HEIC, HEIF)"
            ),
        )
    )
    chunks = [
        fixtures[0:4],
        fixtures[4:8],
        fixtures[8:12],
        fixtures[12:16],
        fixtures[16:20],
        fixtures[20:24],
        fixtures[24:25],
    ]
    plans = {
        _chunk_key(chunks[0]): {
            "kind": "json",
            "body": {
                "results": [
                    _success_result(1, 0),
                    _error_result(1, error_type="BatchProcessingError", error="could not parse"),
                    _unusable_success_result(2),
                    _error_result(
                        3,
                        error_type="DocumentSizeGuardError",
                        error="document exceeded size guard",
                    ),
                ]
            },
            "delay": 0.05,
            "name": "chunk-1",
        },
        _chunk_key(chunks[1]): {
            "kind": "json",
            "status": 503,
            "body": {"detail": "upstream unavailable"},
            "delay": 0.04,
            "name": "chunk-2",
        },
        _chunk_key(chunks[2]): {
            "kind": "json",
            "body": {"detail": "missing results"},
            "delay": 0.03,
            "name": "chunk-3",
        },
        _chunk_key(chunks[3]): {
            "kind": "json",
            "body": {"results": []},
            "delay": 0.02,
            "name": "chunk-4",
        },
        _chunk_key(chunks[4]): {
            "kind": "invalid_json",
            "text": "not-json",
            "delay": 0.01,
            "name": "chunk-5",
        },
        _chunk_key(chunks[5]): {
            "kind": "timeout",
            "name": "chunk-6",
        },
        _chunk_key(chunks[6]): {
            "kind": "request_error",
            "name": "chunk-7",
        },
    }

    _install_fake_make_client(monkeypatch, plans)
    sequential_rows, sequential_chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    _install_fake_make_client(monkeypatch, plans)
    concurrent_rows, concurrent_chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=3,
    )

    assert sequential_chunk_count == 7
    assert concurrent_chunk_count == 7
    assert _row_signatures(sequential_rows) == _row_signatures(concurrent_rows)
    assert [row.metadata["failure_tag"] for row in concurrent_rows] == [
        None,
        "result_error",
        "unusable_result",
        "expected_warning_result",
        "http_503",
        "http_503",
        "http_503",
        "http_503",
        "missing_results_array",
        "missing_results_array",
        "missing_results_array",
        "missing_results_array",
        "missing_result",
        "missing_result",
        "missing_result",
        "missing_result",
        "invalid_json_response",
        "invalid_json_response",
        "invalid_json_response",
        "invalid_json_response",
        "request_timeout",
        "request_timeout",
        "request_timeout",
        "request_timeout",
        "request_error",
        "unsupported_fixture",
    ]


def test_token_expiry_refreshes_client_and_retries_chunk_once(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: [
            {
                "kind": "json",
                "status": 401,
                "body": {"detail": "OpenID Connect token expired: JWT has expired"},
                "name": "expired-token",
            },
            {
                "kind": "json",
                "body": {"results": [_success_result(1, 0)]},
                "name": "success",
            },
        ]
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert chunk_count == 1
    assert attempts[key] == 2
    assert state["client_count"] == 2
    assert state["force_refresh_count"] == 0
    assert rows[0].metadata["ok"] is True
    assert rows[0].metadata["failure_tag"] is None
    assert rows[0].metadata["batch_attempt_count"] == 2
    assert rows[0].metadata["batch_retry_reason"] == "token_expired"
    assert rows[0].metadata["batch_final_attempt_error_type"] is None


def test_run_export_manifest_records_retry_policy_and_summary(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    registry_path = _write_fixture_registry(
        tmp_path / "fixture_registry.yaml",
        file_type_counts={"ACR": 1},
    )
    _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    key = _registry_chunk_key("ACR", range(1, 2))
    plans = {
        key: [
            {
                "kind": "json",
                "status": 401,
                "body": {"detail": "OpenID Connect token expired: JWT has expired"},
                "name": "expired-token",
            },
            {
                "kind": "json",
                "body": _success_body_for_key(key),
                "name": "success",
            },
        ]
    }
    _install_fake_make_client(monkeypatch, plans)

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=tmp_path / "output",
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=1,
        max_concurrent_file_types=1,
        token_expiry_retries=1,
        transient_chunk_retries=1,
        rate_limit_retries=1,
        rate_limit_backoff_secs=2.0,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert manifest["token_expiry_retries"] == 1
    assert manifest["transient_chunk_retries"] == 1
    assert manifest["rate_limit_retries"] == 1
    assert manifest["rate_limit_backoff_secs"] == 2.0
    assert manifest["file_types"]["ACR"]["retry_summary"] == {
        "rows_with_retries": 1,
        "max_batch_attempt_count": 2,
        "batch_retry_reasons": {"token_expired": 1},
        "batch_final_attempt_error_types": {},
    }


def test_persistent_token_expiry_after_refresh_is_recorded_clearly(monkeypatch):
    fixtures = [_make_fixture(record_id) for record_id in range(1, 3)]
    key = _chunk_key(fixtures)
    token_body = {"detail": "OpenID Connect token expired: JWT has expired"}
    plans = {
        key: [
            {"kind": "json", "status": 401, "body": token_body, "name": "expired-1"},
            {"kind": "json", "status": 401, "body": token_body, "name": "expired-2"},
        ]
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert attempts[key] == 2
    assert state["force_refresh_count"] == 0
    assert [row.metadata["ok"] for row in rows] == [False, False]
    assert {row.metadata["failure_tag"] for row in rows} == {"persistent_token_expired"}
    assert {row.metadata["error_type"] for row in rows} == {"TokenExpired"}
    assert {row.metadata["batch_http_status"] for row in rows} == {401}
    assert {row.metadata["batch_attempt_count"] for row in rows} == {2}
    assert {row.metadata["batch_retry_reason"] for row in rows} == {"token_expired"}
    assert {row.metadata["batch_final_attempt_error_type"] for row in rows} == {"TokenExpired"}
    assert all("JWT has expired" in str(row.metadata["error"]) for row in rows)


def test_read_timeout_retries_once_and_succeeds_without_refreshing_client(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: [
            {"kind": "timeout", "name": "timeout"},
            {"kind": "json", "body": {"results": [_success_result(1, 0)]}, "name": "success"},
        ]
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert attempts[key] == 2
    assert state["client_count"] == 1
    assert state["force_refresh_count"] == 0
    assert rows[0].metadata["ok"] is True
    assert rows[0].metadata["batch_attempt_count"] == 2
    assert rows[0].metadata["batch_retry_reason"] == "read_timeout"


def test_remote_protocol_error_retries_once_and_succeeds_without_refreshing_client(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: [
            {"kind": "remote_protocol", "name": "disconnect"},
            {"kind": "json", "body": {"results": [_success_result(1, 0)]}, "name": "success"},
        ]
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert attempts[key] == 2
    assert state["client_count"] == 1
    assert state["force_refresh_count"] == 0
    assert rows[0].metadata["ok"] is True
    assert rows[0].metadata["batch_attempt_count"] == 2
    assert rows[0].metadata["batch_retry_reason"] == "remote_protocol_error"


def test_rate_limit_retry_respects_retry_after_and_succeeds(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: [
            {
                "kind": "json",
                "status": 429,
                "headers": {"Retry-After": "3"},
                "body": {"error": "Rate limit exceeded"},
                "name": "rate-limited",
            },
            {"kind": "json", "body": {"results": [_success_result(1, 0)]}, "name": "success"},
        ]
    }
    slept: list[float] = []
    monkeypatch.setattr(batch_ground_truth_workflow, "_sleep", slept.append)
    state = _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert attempts[key] == 2
    assert slept == [3.0]
    assert rows[0].metadata["ok"] is True
    assert rows[0].metadata["failure_tag"] is None
    assert rows[0].metadata["batch_attempt_count"] == 2
    assert rows[0].metadata["batch_retry_reason"] == "rate_limited"
    assert rows[0].metadata["batch_final_attempt_error_type"] is None
    assert is_clean_candidate(rows[0]) is True


def test_rate_limit_retry_exhaustion_is_triaged_as_rate_limited(monkeypatch):
    fixtures = [_make_fixture(record_id) for record_id in range(1, 3)]
    key = _chunk_key(fixtures)
    plans = {
        key: [
            {
                "kind": "json",
                "status": 429,
                "body": {"error": "Rate limit exceeded"},
                "name": "rate-limited-1",
            },
            {
                "kind": "json",
                "status": 429,
                "body": {"error": "Rate limit exceeded"},
                "name": "rate-limited-2",
            },
        ]
    }
    slept: list[float] = []
    monkeypatch.setattr(batch_ground_truth_workflow, "_sleep", slept.append)
    state = _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
        rate_limit_retries=1,
        rate_limit_backoff_secs=0.25,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert attempts[key] == 2
    assert slept == [0.25]
    assert {row.metadata["ok"] for row in rows} == {False}
    assert {row.metadata["failure_tag"] for row in rows} == {"http_429"}
    assert {row.metadata["batch_http_status"] for row in rows} == {429}
    assert {row.metadata["batch_attempt_count"] for row in rows} == {2}
    assert {row.metadata["batch_retry_reason"] for row in rows} == {"rate_limited"}
    assert {row.metadata["batch_final_attempt_error_type"] for row in rows} == {"RateLimited"}
    assert all(is_clean_candidate(row) is False for row in rows)

    triage_row = build_recovery_triage_row(rows[0])
    assert triage_row["recovery_class"] == "rate_limited"
    assert triage_row["gt_candidate_status"] == "rerun_candidate"
    assert triage_row["recovery_action"] == "targeted_rerun_with_lower_concurrency_or_backoff"


@pytest.mark.parametrize(
    ("kind", "failure_tag", "error_type", "retry_reason"),
    [
        ("timeout", "request_timeout", "ReadTimeout", "read_timeout"),
        ("remote_protocol", "request_error", "RemoteProtocolError", "remote_protocol_error"),
    ],
)
def test_persistent_transient_chunk_failures_are_recorded_after_retry(
    monkeypatch,
    kind,
    failure_tag,
    error_type,
    retry_reason,
):
    fixtures = [_make_fixture(record_id) for record_id in range(1, 3)]
    key = _chunk_key(fixtures)
    plans = {
        key: [
            {"kind": kind, "name": "first-failure"},
            {"kind": kind, "name": "second-failure"},
        ]
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert attempts[key] == 2
    assert {row.metadata["ok"] for row in rows} == {False}
    assert {row.metadata["failure_tag"] for row in rows} == {failure_tag}
    assert {row.metadata["error_type"] for row in rows} == {error_type}
    assert {row.metadata["batch_attempt_count"] for row in rows} == {2}
    assert {row.metadata["batch_retry_reason"] for row in rows} == {retry_reason}
    assert {row.metadata["batch_final_attempt_error_type"] for row in rows} == {error_type}


def test_expected_row_level_api_errors_are_not_retried(monkeypatch):
    fixtures = [
        _make_fixture(
            1,
            batch_expected_error_type="DocumentSizeGuardError",
            batch_expected_warning="document exceeded size guard",
        )
    ]
    key = _chunk_key(fixtures)
    plans = {
        key: {
            "kind": "json",
            "body": {
                "results": [
                    _error_result(
                        0,
                        error_type="DocumentSizeGuardError",
                        error="document exceeded size guard",
                    )
                ]
            },
        }
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert attempts[key] == 1
    assert rows[0].metadata["failure_tag"] == "expected_warning_result"
    assert rows[0].metadata["error_type"] == "DocumentSizeGuardError"
    assert rows[0].metadata["batch_attempt_count"] == 1
    assert rows[0].metadata["batch_retry_reason"] is None


def test_unusable_result_from_200_response_is_not_retried(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: {
            "kind": "json",
            "body": {"results": [_unusable_success_result(0)]},
        }
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    attempts = state["attempts"]
    assert isinstance(attempts, dict)
    assert attempts[key] == 1
    assert rows[0].metadata["failure_tag"] == "unusable_result"
    assert rows[0].metadata["error_type"] == "ValueError"
    assert rows[0].metadata["batch_attempt_count"] == 1
    assert rows[0].metadata["batch_retry_reason"] is None


def test_clean_candidate_classification_accepts_usable_success_with_null_fields(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: {
            "kind": "json",
            "body": {"results": [_null_field_success_result(1, 0)]},
        }
    }
    _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    assert is_clean_candidate(rows[0]) is True
    assert classify_export_row(rows[0]) == {
        "recovery_class": "clean_success",
        "recovery_action": "include_in_clean_gt",
        "gt_candidate_status": "clean_candidate",
    }


@pytest.mark.parametrize(
    ("plans_for_key", "retry_kwargs"),
    [
        (
            [
                {
                    "kind": "json",
                    "status": 401,
                    "body": {"detail": "OpenID Connect token expired: JWT has expired"},
                    "name": "expired-token",
                }
            ],
            {"token_expiry_retries": 0},
        ),
        ([{"kind": "timeout", "name": "timeout"}], {"transient_chunk_retries": 0}),
        ([{"kind": "remote_protocol", "name": "disconnect"}], {"transient_chunk_retries": 0}),
    ],
)
def test_recovery_triage_classifies_transient_or_auth_failures(
    monkeypatch,
    plans_for_key,
    retry_kwargs,
):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {key: plans_for_key}
    _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
        **retry_kwargs,
    )

    triage_row = build_recovery_triage_row(rows[0])
    assert triage_row["recovery_class"] == "transient_or_auth_failure"
    assert triage_row["gt_candidate_status"] == "rerun_candidate"
    assert triage_row["recovery_action"] == "targeted_rerun_after_retry_or_token_fix"


def test_recovery_triage_classifies_document_size_guard_as_not_current_gt(monkeypatch):
    fixtures = [
        _make_fixture(
            1,
            batch_expected_error_type="DocumentSizeGuardError",
            batch_expected_warning="document exceeded size guard",
        )
    ]
    key = _chunk_key(fixtures)
    plans = {
        key: {
            "kind": "json",
            "body": {
                "results": [
                    _error_result(
                        0,
                        error_type="DocumentSizeGuardError",
                        error="document exceeded size guard",
                    )
                ]
            },
        }
    }
    _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    triage_row = build_recovery_triage_row(rows[0])
    assert triage_row["recovery_class"] == "document_size_guard"
    assert triage_row["gt_candidate_status"] == "not_gt_candidate_currently"
    assert triage_row["recovery_action"] == "exclude_from_clean_gt_or_replace_fixture"


def test_recovery_triage_classifies_unsupported_fixture():
    row = batch_ground_truth_workflow._skipped_export_row(
        _make_fixture(
            1,
            include_in_batch=False,
            skip_reason=(
                "unsupported file extension '.xlsx' "
                "(supported: PDF, PNG, JPG, JPEG, TIFF, TIF, HEIC, HEIF)"
            ),
        ),
        output_generated_at="2026-04-24T00:00:00+00:00",
    )

    triage_row = build_recovery_triage_row(row)
    assert triage_row["recovery_class"] == "unsupported_fixture"
    assert triage_row["gt_candidate_status"] == "fixture_review_required"


def test_recovery_triage_classifies_multi_account_document(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: {
            "kind": "json",
            "body": {
                "results": [
                    _error_result(
                        0,
                        error_type="MultiAccountDocumentError",
                        error="multiple accounts detected",
                    )
                ]
            },
        }
    }
    _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    triage_row = build_recovery_triage_row(rows[0])
    assert triage_row["recovery_class"] == "multi_account_document"
    assert triage_row["gt_candidate_status"] == "fixture_review_required"


def test_recovery_triage_classifies_http_200_no_payload_quality_gate(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: {
            "kind": "json",
            "body": {"results": [_quality_gated_no_payload_result(0)]},
        }
    }
    _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    triage_row = build_recovery_triage_row(rows[0])
    assert rows[0].metadata["failure_tag"] == "unusable_result"
    assert triage_row["recovery_class"] == "http_200_no_payload_quality_gate"
    assert triage_row["gt_candidate_status"] == "not_gt_candidate_currently"
    assert triage_row["extractionStatus"] == "not_attempted"
    assert triage_row["documentQuality"] == "Failed: failed document_classification"
    assert triage_row["qualityCheck.issueDescription"] == "Failed: failed document_classification"
    assert triage_row["qualityScore"] == 0.0
    assert triage_row["summaryResult_count"] == 0
    assert triage_row["summaryOCR_count"] == 0
    assert triage_row["calculatedFields_count"] == 0
    assert triage_row["transactionsOCR_count"] == 0


def test_recovery_triage_classifies_http_200_no_payload_unknown(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: {
            "kind": "json",
            "body": {"results": [_unknown_no_payload_result(0)]},
        }
    }
    _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    triage_row = build_recovery_triage_row(rows[0])
    assert rows[0].metadata["failure_tag"] == "unusable_result"
    assert triage_row["recovery_class"] == "http_200_no_payload_unknown"
    assert triage_row["gt_candidate_status"] == "api_behavior_review_required"


def test_recovery_triage_keeps_generic_unusable_shape_separate(monkeypatch):
    fixtures = [_make_fixture(1)]
    key = _chunk_key(fixtures)
    plans = {
        key: {
            "kind": "json",
            "body": {"results": [_unusable_success_result(0)]},
        }
    }
    _install_fake_make_client(monkeypatch, plans)

    rows, _chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=1,
    )

    triage_row = build_recovery_triage_row(rows[0])
    assert rows[0].metadata["failure_tag"] == "unusable_result"
    assert triage_row["recovery_class"] == "malformed_or_unexpected_response_shape"
    assert triage_row["gt_candidate_status"] == "api_behavior_review_required"


def test_concurrent_retry_keeps_final_rows_in_source_order(monkeypatch):
    fixtures = [_make_fixture(record_id) for record_id in range(1, 9)]
    chunk_one = fixtures[:4]
    chunk_two = fixtures[4:]
    plans = {
        _chunk_key(chunk_one): [
            {"kind": "timeout", "name": "chunk-1-timeout"},
            {
                "kind": "json",
                "body": {
                    "results": [
                        _success_result(record_id, index)
                        for index, record_id in enumerate(range(1, 5))
                    ]
                },
                "delay": 0.03,
                "name": "chunk-1-success",
            },
        ],
        _chunk_key(chunk_two): {
            "kind": "json",
            "body": {
                "results": [
                    _success_result(record_id, index)
                    for index, record_id in enumerate(range(5, 9))
                ]
            },
            "name": "chunk-2",
        },
    }
    state = _install_fake_make_client(monkeypatch, plans)

    rows, chunk_count = batch_ground_truth_workflow._execute_file_type(
        fixtures=fixtures,
        output_generated_at="2026-04-24T00:00:00+00:00",
        max_concurrent_chunks=2,
    )

    assert chunk_count == 2
    assert state["max_active"] == 2
    assert [row.metadata["source_row"] for row in rows] == [fixture.source_row for fixture in fixtures]
    assert [row.metadata["batch_chunk_number"] for row in rows] == [1, 1, 1, 1, 2, 2, 2, 2]
    assert [row.metadata["batch_attempt_count"] for row in rows] == [2, 2, 2, 2, 1, 1, 1, 1]


def test_iap_token_cache_uses_lock_for_concurrent_mint_and_forced_refresh(monkeypatch):
    from tests import client as client_module

    client_module.clear_iap_token_cache()
    mint_count = 0
    mint_lock = threading.Lock()

    def fake_fetch_iap_id_token() -> str:
        nonlocal mint_count
        time.sleep(0.02)
        with mint_lock:
            mint_count += 1
            return f"token-{mint_count}"

    monkeypatch.setattr(client_module, "_fetch_iap_id_token", fake_fetch_iap_id_token)
    worker_count = 8
    barrier = threading.Barrier(worker_count)

    def get_token_after_barrier() -> str:
        barrier.wait(timeout=5.0)
        return client_module.get_iap_bearer()

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        tokens = list(executor.map(lambda _index: get_token_after_barrier(), range(worker_count)))

    assert tokens == ["token-1"] * worker_count
    assert mint_count == 1
    assert client_module.clear_iap_token_cache(expired_token="not-token-1") is False
    assert client_module.get_iap_bearer() == "token-1"
    assert mint_count == 1
    assert client_module.clear_iap_token_cache(expired_token="token-1") is True
    assert client_module.get_iap_bearer() == "token-2"
    assert mint_count == 2
    assert client_module.get_iap_bearer(force_refresh=True) == "token-3"
    assert client_module.get_iap_bearer() == "token-3"
    assert mint_count == 3
    client_module.clear_iap_token_cache()


def test_run_export_preserves_file_type_sequential_default(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    registry_path = _write_fixture_registry(
        tmp_path / "fixture_registry.yaml",
        file_type_counts={"ACR": 1, "TIN": 1},
    )
    output_dir = tmp_path / "output"
    artifact_run_dir = _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    acr_key = _registry_chunk_key("ACR", range(1, 2))
    tin_key = _registry_chunk_key("TIN", range(1, 2))
    plans = {
        acr_key: {
            "kind": "json",
            "body": _success_body_for_key(acr_key),
            "delay": 0.03,
            "name": "ACR",
        },
        tin_key: {
            "kind": "json",
            "body": _success_body_for_key(tin_key),
            "name": "TIN",
        },
    }
    state = _install_fake_make_client(monkeypatch, plans)

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=output_dir,
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=1,
        max_concurrent_file_types=1,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert state["client_count"] == 2
    assert state["max_active"] == 1
    assert state["completion_order"] == ["ACR", "TIN"]
    assert result.batch_artifact_run_dir == artifact_run_dir
    assert result.selected_file_types == ("ACR", "TIN")
    assert [file_type_result.file_type for file_type_result in result.file_type_results] == [
        "ACR",
        "TIN",
    ]
    assert manifest["max_concurrent_file_types"] == 1
    assert manifest["max_concurrent_chunks"] == 1
    assert manifest["effective_max_concurrent_batch_requests"] == 1
    assert manifest["selected_file_types"] == ["ACR", "TIN"]


def test_run_export_runs_file_types_concurrently_and_keeps_manifest_order(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    registry_path = _write_fixture_registry(
        tmp_path / "fixture_registry.yaml",
        file_type_counts={"ACR": 1, "TIN": 1},
    )
    output_dir = tmp_path / "output"
    _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    acr_key = _registry_chunk_key("ACR", range(1, 2))
    tin_key = _registry_chunk_key("TIN", range(1, 2))
    barrier = threading.Barrier(2)
    plans = {
        acr_key: {
            "kind": "json",
            "body": _success_body_for_key(acr_key),
            "barrier": barrier,
            "delay": 0.03,
            "name": "ACR",
        },
        tin_key: {
            "kind": "json",
            "body": _success_body_for_key(tin_key),
            "barrier": barrier,
            "name": "TIN",
        },
    }
    state = _install_fake_make_client(monkeypatch, plans)

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=output_dir,
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=1,
        max_concurrent_file_types=2,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert state["max_active"] == 2
    assert state["completion_order"] == ["TIN", "ACR"]
    assert result.selected_file_types == ("ACR", "TIN")
    assert [file_type_result.file_type for file_type_result in result.file_type_results] == [
        "ACR",
        "TIN",
    ]
    assert manifest["selected_file_types"] == ["ACR", "TIN"]
    assert list(manifest["file_types"]) == ["ACR", "TIN"]
    assert manifest["max_concurrent_file_types"] == 2
    assert manifest["max_concurrent_chunks"] == 1
    assert manifest["effective_max_concurrent_batch_requests"] == 2


def test_run_export_keeps_workbook_paths_associated_with_file_types(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    registry_path = _write_fixture_registry(
        tmp_path / "fixture_registry.yaml",
        file_type_counts={"ACR": 1, "TIN": 1},
    )
    output_dir = tmp_path / "output"
    _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    acr_key = _registry_chunk_key("ACR", range(1, 2))
    tin_key = _registry_chunk_key("TIN", range(1, 2))
    plans = {
        acr_key: {"kind": "json", "body": _success_body_for_key(acr_key), "name": "ACR"},
        tin_key: {"kind": "json", "body": _success_body_for_key(tin_key), "name": "TIN"},
    }
    _install_fake_make_client(monkeypatch, plans)

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=output_dir,
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=1,
        max_concurrent_file_types=2,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    result_by_type = {
        file_type_result.file_type: file_type_result
        for file_type_result in result.file_type_results
    }
    for file_type in ("ACR", "TIN"):
        workbook_path = result_by_type[file_type].workbook_path
        assert workbook_path == Path(manifest["file_types"][file_type]["workbook_path"])
        assert workbook_path.exists()
        workbook = load_workbook(workbook_path, data_only=True)
        assert workbook.sheetnames == [file_type, "_meta"]
        meta_sheet = workbook["_meta"]
        headers = [meta_sheet.cell(1, column).value for column in range(1, meta_sheet.max_column + 1)]
        normalized_file_type_column = headers.index("normalized_file_type") + 1
        assert meta_sheet.cell(2, normalized_file_type_column).value == file_type


def test_run_export_writes_clean_workbook_manifest_and_recovery_triage(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    registry_path = _write_mixed_fixture_registry(tmp_path / "fixture_registry.yaml")
    output_dir = tmp_path / "output"
    _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    key = (
        "gs://bucket/clean.pdf",
        "gs://bucket/quality-gated.pdf",
        "gs://bucket/size-guard.pdf",
    )
    plans = {
        key: {
            "kind": "json",
            "body": {
                "results": [
                    _success_result(1, 0),
                    _quality_gated_no_payload_result(1),
                    _error_result(
                        2,
                        error_type="DocumentSizeGuardError",
                        error="document exceeded size guard",
                    ),
                ]
            },
        }
    }
    _install_fake_make_client(monkeypatch, plans)

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=output_dir,
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=1,
        max_concurrent_file_types=1,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    clean_manifest = json.loads(result.clean_manifest_path.read_text(encoding="utf-8"))
    recovery_rows = json.loads(result.recovery_triage_json_path.read_text(encoding="utf-8"))
    csv_lines = result.recovery_triage_csv_path.read_text(encoding="utf-8").splitlines()
    assert len(recovery_rows) == 3
    assert len(csv_lines) == 4
    assert {row["recovery_class"] for row in recovery_rows} == {
        "http_200_no_payload_quality_gate",
        "document_size_guard",
        "unsupported_fixture",
    }
    assert clean_manifest["total_source_rows"] == 4
    assert clean_manifest["clean_included_rows"] == 1
    assert clean_manifest["triaged_rejected_rows"] == 3
    assert clean_manifest["clean_rows_by_fileType"] == {"Payslip": 1}
    assert clean_manifest["triaged_rows_by_fileType"] == {"Payslip": 3}
    assert clean_manifest["triaged_rows_by_recovery_class"] == {
        "http_200_no_payload_quality_gate": 1,
        "document_size_guard": 1,
        "unsupported_fixture": 1,
    }
    assert clean_manifest["triaged_rows_by_gt_candidate_status"] == {
        "not_gt_candidate_currently": 2,
        "fixture_review_required": 1,
    }

    file_type_result = result.file_type_results[0]
    assert file_type_result.clean_rows == 1
    assert file_type_result.triaged_rows == 3
    assert file_type_result.clean_workbook_path == (
        output_dir / "clean_workbooks" / clean_workbook_filename_for("Payslip")
    )
    assert manifest["file_types"]["Payslip"]["workbook_path"] == str(file_type_result.workbook_path)
    assert manifest["file_types"]["Payslip"]["clean_workbook_path"] == str(file_type_result.clean_workbook_path)

    audit_workbook = load_workbook(file_type_result.workbook_path, data_only=True)
    audit_meta_sheet = audit_workbook["_meta"]
    assert audit_meta_sheet.max_row == 5

    clean_workbook = load_workbook(file_type_result.clean_workbook_path, data_only=True)
    clean_sheet = clean_workbook["Payslip"]
    clean_headers = [
        clean_sheet.cell(1, column).value
        for column in range(1, clean_sheet.max_column + 1)
    ]
    parse_success_column = clean_headers.index("parse_success") + 1
    error_column = clean_headers.index("error") + 1
    assert clean_sheet.max_row == 2
    assert clean_sheet.cell(2, parse_success_column).value is True
    assert clean_sheet.cell(2, error_column).value is None
    assert all(
        clean_sheet.cell(row, parse_success_column).value is not False
        for row in range(2, clean_sheet.max_row + 1)
    )


@pytest.mark.parametrize(
    (
        "gt_skip_reason",
        "gt_classification",
        "gt_recovery_action",
        "batch_error_type",
        "expected_failure_tag",
        "expected_recovery_class",
        "expected_gt_status",
    ),
    [
        (
            "document_size_guard",
            "fixture_too_large",
            "reduce_fixture",
            "DocumentSizeGuardError",
            "document_size_guard",
            "document_size_guard",
            "not_gt_candidate_currently",
        ),
        (
            "multi_account_document",
            "multi_account_fixture",
            "split_fixture",
            "MultiAccountDocumentError",
            "multi_account_document",
            "multi_account_document",
            "fixture_review_required",
        ),
    ],
)
def test_gt_extraction_excluded_fixture_is_skipped_but_kept_in_audit_and_triage(
    monkeypatch,
    tmp_path,
    gt_skip_reason,
    gt_classification,
    gt_recovery_action,
    batch_error_type,
    expected_failure_tag,
    expected_recovery_class,
    expected_gt_status,
):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    registry_path = _write_gt_exclusion_registry(
        tmp_path / "fixture_registry.yaml",
        excluded_fixture={
            "name": "excluded",
            "file_type": "Payslip",
            "gcs_uri": "gs://bucket/excluded.pdf",
            "source_folder": "Payslip",
            "source_file_type": "Payslip",
            "source_file_type_status": "✓",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "source_row": 11,
            "verification_status": "confirmed",
            "enabled": True,
            "batch_expected_error_type": batch_error_type,
            "batch_expected_error": "fixture cannot produce clean GT as-is",
            "gt_extraction_eligible": False,
            "gt_extraction_skip_reason": gt_skip_reason,
            "gt_extraction_classification": gt_classification,
            "gt_clean_eligible": False,
            "negative_audit_useful": True,
            "gt_recovery_action": gt_recovery_action,
        },
    )
    output_dir = tmp_path / "output"
    _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    key = ("gs://bucket/clean.pdf",)
    _install_fake_make_client(
        monkeypatch,
        {key: {"kind": "json", "body": _success_body_for_key(key)}},
    )

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=output_dir,
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=1,
        max_concurrent_file_types=1,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    clean_manifest = json.loads(result.clean_manifest_path.read_text(encoding="utf-8"))
    recovery_rows = json.loads(result.recovery_triage_json_path.read_text(encoding="utf-8"))

    assert manifest["file_types"]["Payslip"]["executable_rows"] == 1
    assert manifest["file_types"]["Payslip"]["skipped_rows"] == 1
    assert manifest["file_types"]["Payslip"]["chunk_count"] == 1
    assert manifest["skipped_rows"][0]["gt_extraction_skip_reason"] == gt_skip_reason
    assert clean_manifest["clean_included_rows"] == 1
    assert clean_manifest["gt_extraction_excluded_rows"] == 1
    assert clean_manifest["gt_extraction_excluded_rows_by_fileType"] == {"Payslip": 1}
    assert len(recovery_rows) == 1
    triage_row = recovery_rows[0]
    assert triage_row["failure_tag"] == expected_failure_tag
    assert triage_row["error_type"] == batch_error_type
    assert triage_row["recovery_class"] == expected_recovery_class
    assert triage_row["gt_candidate_status"] == expected_gt_status
    assert triage_row["gt_extraction_excluded"] is True
    assert triage_row["gt_extraction_classification"] == gt_classification

    audit_workbook = load_workbook(result.file_type_results[0].workbook_path, data_only=True)
    assert audit_workbook["_meta"].max_row == 3
    clean_workbook = load_workbook(result.file_type_results[0].clean_workbook_path, data_only=True)
    assert clean_workbook["Payslip"].max_row == 2


def test_quality_gate_rows_are_future_skipped_only_when_explicitly_tagged(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    fixtures = [
        {
            "name": "quality-tagged",
            "file_type": "Payslip",
            "gcs_uri": "gs://bucket/quality-tagged.pdf",
            "source_folder": "Payslip",
            "source_file_type": "Payslip",
            "source_file_type_status": "✓",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "source_row": 10,
            "verification_status": "confirmed",
            "enabled": True,
            "gt_extraction_eligible": False,
            "gt_extraction_skip_reason": "quality_gate_no_payload",
            "gt_extraction_classification": "fixture_quality_gate_failed",
            "gt_clean_eligible": False,
            "negative_audit_useful": True,
            "gt_recovery_action": "review_fixture_quality",
        },
        {
            "name": "quality-untagged",
            "file_type": "Payslip",
            "gcs_uri": "gs://bucket/quality-untagged.pdf",
            "source_folder": "Payslip",
            "source_file_type": "Payslip",
            "source_file_type_status": "✓",
            "source_assignee": "Thor",
            "source_workflow_status": "Pending",
            "source_row": 11,
            "verification_status": "confirmed",
            "enabled": True,
        },
    ]
    registry_path = tmp_path / "fixture_registry.yaml"
    write_registry_document(
        {
            "schema_version": 2,
            "total": len(fixtures),
            "composite_rows_split": 0,
            "counts": {"confirmed": len(fixtures)},
            "fixtures": fixtures,
        },
        output_paths=(registry_path,),
    )
    output_dir = tmp_path / "output"
    _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    key = ("gs://bucket/quality-untagged.pdf",)
    _install_fake_make_client(
        monkeypatch,
        {key: {"kind": "json", "body": {"results": [_quality_gated_no_payload_result(0)]}}},
    )

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=output_dir,
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=1,
        max_concurrent_file_types=1,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    recovery_rows = json.loads(result.recovery_triage_json_path.read_text(encoding="utf-8"))
    assert manifest["file_types"]["Payslip"]["executable_rows"] == 1
    assert manifest["file_types"]["Payslip"]["skipped_rows"] == 1
    assert [row["source_basename"] for row in recovery_rows] == [
        "quality-tagged",
        "quality-untagged",
    ]
    assert [row["recovery_class"] for row in recovery_rows] == [
        "http_200_no_payload_quality_gate",
        "http_200_no_payload_quality_gate",
    ]
    assert recovery_rows[0]["gt_extraction_excluded"] is True
    assert recovery_rows[0]["batch_http_status"] is None
    assert recovery_rows[1]["gt_extraction_excluded"] is False
    assert recovery_rows[1]["batch_http_status"] == 200
    assert result.file_type_results[0].clean_rows == 0


def test_run_export_excludes_rate_limited_rows_from_clean_and_triages_them(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    registry_path = _write_fixture_registry(
        tmp_path / "fixture_registry.yaml",
        file_type_counts={"ACR": 5},
    )
    output_dir = tmp_path / "output"
    _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    first_chunk = _registry_chunk_key("ACR", range(1, 5))
    second_chunk = _registry_chunk_key("ACR", range(5, 6))
    plans = {
        first_chunk: {"kind": "json", "body": _success_body_for_key(first_chunk)},
        second_chunk: {
            "kind": "json",
            "status": 429,
            "body": {"error": "Rate limit exceeded"},
        },
    }
    _install_fake_make_client(monkeypatch, plans)

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=output_dir,
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=1,
        max_concurrent_file_types=1,
        rate_limit_retries=0,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    clean_manifest = json.loads(result.clean_manifest_path.read_text(encoding="utf-8"))
    recovery_rows = json.loads(result.recovery_triage_json_path.read_text(encoding="utf-8"))
    with result.recovery_triage_csv_path.open(newline="", encoding="utf-8") as handle:
        csv_rows = list(csv.DictReader(handle))
    assert len(recovery_rows) == 1
    assert len(csv_rows) == 1
    assert recovery_rows[0]["failure_tag"] == "http_429"
    assert recovery_rows[0]["batch_http_status"] == 429
    assert recovery_rows[0]["recovery_class"] == "rate_limited"
    assert recovery_rows[0]["gt_candidate_status"] == "rerun_candidate"
    assert recovery_rows[0]["recovery_action"] == "targeted_rerun_with_lower_concurrency_or_backoff"
    assert clean_manifest["clean_included_rows"] == 4
    assert clean_manifest["triaged_rejected_rows"] == 1
    assert clean_manifest["triaged_rows_by_recovery_class"] == {"rate_limited": 1}
    assert clean_manifest["triaged_rows_by_gt_candidate_status"] == {"rerun_candidate": 1}
    assert manifest["rate_limit_retries"] == 0
    assert manifest["file_types"]["ACR"]["retry_summary"] == {
        "rows_with_retries": 0,
        "max_batch_attempt_count": 1,
        "batch_retry_reasons": {},
        "batch_final_attempt_error_types": {"RateLimited": 1},
    }

    file_type_result = result.file_type_results[0]
    assert file_type_result.clean_rows == 4
    assert file_type_result.triaged_rows == 1
    clean_workbook = load_workbook(file_type_result.clean_workbook_path, data_only=True)
    clean_sheet = clean_workbook["ACR"]
    clean_headers = [
        clean_sheet.cell(1, column).value
        for column in range(1, clean_sheet.max_column + 1)
    ]
    parse_success_column = clean_headers.index("parse_success") + 1
    error_column = clean_headers.index("error") + 1
    assert clean_sheet.max_row == 5
    for row in range(2, clean_sheet.max_row + 1):
        assert clean_sheet.cell(row, parse_success_column).value is True
        assert clean_sheet.cell(row, error_column).value is None


def test_run_export_keeps_chunk_numbers_stable_under_file_type_concurrency(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    registry_path = _write_fixture_registry(
        tmp_path / "fixture_registry.yaml",
        file_type_counts={"ACR": 5, "TIN": 5},
    )
    output_dir = tmp_path / "output"
    _configure_batch_artifact_dir(monkeypatch, tmp_path)
    layout = load_reference_template(reference_path)
    acr_chunk_one = _registry_chunk_key("ACR", range(1, 5))
    acr_chunk_two = _registry_chunk_key("ACR", range(5, 6))
    tin_chunk_one = _registry_chunk_key("TIN", range(1, 5))
    tin_chunk_two = _registry_chunk_key("TIN", range(5, 6))
    barrier = threading.Barrier(4)
    plans = {
        acr_chunk_one: {
            "kind": "json",
            "body": _success_body_for_key(acr_chunk_one),
            "barrier": barrier,
            "delay": 0.03,
            "name": "ACR-1",
        },
        acr_chunk_two: {
            "kind": "json",
            "body": _success_body_for_key(acr_chunk_two),
            "barrier": barrier,
            "name": "ACR-2",
        },
        tin_chunk_one: {
            "kind": "json",
            "body": _success_body_for_key(tin_chunk_one),
            "barrier": barrier,
            "delay": 0.03,
            "name": "TIN-1",
        },
        tin_chunk_two: {
            "kind": "json",
            "body": _success_body_for_key(tin_chunk_two),
            "barrier": barrier,
            "name": "TIN-2",
        },
    }
    state = _install_fake_make_client(monkeypatch, plans)

    result = batch_ground_truth_workflow.run_batch_ground_truth_export(
        fixture_registry=registry_path,
        reference_workbook=reference_path,
        output_dir=output_dir,
        selected_file_types=None,
        template_layout=layout,
        max_concurrent_chunks=2,
        max_concurrent_file_types=2,
    )

    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))
    assert state["max_active"] == 4
    assert manifest["effective_max_concurrent_batch_requests"] == 4
    for file_type_result in result.file_type_results:
        assert file_type_result.chunk_count == 2
        assert manifest["file_types"][file_type_result.file_type]["chunk_count"] == 2
        workbook = load_workbook(file_type_result.workbook_path, data_only=True)
        meta_sheet = workbook["_meta"]
        headers = [meta_sheet.cell(1, column).value for column in range(1, meta_sheet.max_column + 1)]
        chunk_column = headers.index("batch_chunk_number") + 1
        assert [
            meta_sheet.cell(row, chunk_column).value
            for row in range(2, meta_sheet.max_row + 1)
        ] == [1, 1, 1, 1, 2]


@pytest.mark.parametrize("value", ["0", "-1", "not-an-int"])
def test_build_parser_rejects_invalid_max_concurrent_chunks(value):
    parser = export_batch_ground_truth.build_parser()

    with pytest.raises(SystemExit) as exc:
        parser.parse_args(
            [
                "--reference-workbook",
                "reference.xlsx",
                "--max-concurrent-chunks",
                value,
            ]
        )

    assert exc.value.code == 2


@pytest.mark.parametrize("value", ["0", "-1", "not-an-int"])
def test_build_parser_rejects_invalid_max_concurrent_file_types(value):
    parser = export_batch_ground_truth.build_parser()

    with pytest.raises(SystemExit) as exc:
        parser.parse_args(
            [
                "--reference-workbook",
                "reference.xlsx",
                "--max-concurrent-file-types",
                value,
            ]
        )

    assert exc.value.code == 2


@pytest.mark.parametrize(
    "flag",
    ["--token-expiry-retries", "--transient-chunk-retries", "--rate-limit-retries"],
)
@pytest.mark.parametrize("value", ["-1", "not-an-int"])
def test_build_parser_rejects_invalid_retry_counts(flag, value):
    parser = export_batch_ground_truth.build_parser()

    with pytest.raises(SystemExit) as exc:
        parser.parse_args(
            [
                "--reference-workbook",
                "reference.xlsx",
                flag,
                value,
            ]
        )

    assert exc.value.code == 2


@pytest.mark.parametrize("value", ["-1", "not-a-number"])
def test_build_parser_rejects_invalid_rate_limit_backoff(value):
    parser = export_batch_ground_truth.build_parser()

    with pytest.raises(SystemExit) as exc:
        parser.parse_args(
            [
                "--reference-workbook",
                "reference.xlsx",
                "--rate-limit-backoff-secs",
                value,
            ]
        )

    assert exc.value.code == 2


def test_main_passes_concurrency_values_to_workflow(monkeypatch, tmp_path):
    reference_path = _write_reference_workbook(tmp_path / "reference.xlsx")
    source_path = _write_source_workbook(tmp_path / "source.xlsx")
    registry_path = _write_fixture_registry_from_source_workbook(
        source_path,
        tmp_path / "fixture_registry.yaml",
    )
    captured: dict[str, object] = {}

    monkeypatch.setattr(
        export_batch_ground_truth,
        "plan_file_types",
        lambda **kwargs: (
            SimpleNamespace(source_workbook=source_path.resolve()),
            {},
            [
                FileTypePlan(
                    file_type="Payslip",
                    total_rows=1,
                    executable_rows=1,
                    skipped_rows=0,
                    chunk_count=1,
                )
            ],
        ),
    )
    monkeypatch.setattr(
        export_batch_ground_truth,
        "load_reference_template",
        lambda reference_workbook: {"reference_workbook": reference_workbook},
    )

    def fake_run_batch_ground_truth_export(**kwargs):
        captured.update(kwargs)
        return SimpleNamespace(
            output_dir=tmp_path / "output",
            manifest_path=tmp_path / "output" / "manifest.json",
            batch_artifact_run_dir=tmp_path / "batch_artifacts",
            file_type_results=[],
        )

    monkeypatch.setattr(
        export_batch_ground_truth,
        "run_batch_ground_truth_export",
        fake_run_batch_ground_truth_export,
    )

    exit_code = export_batch_ground_truth.main(
        [
            "--fixture-registry",
            str(registry_path),
            "--reference-workbook",
            str(reference_path),
            "--max-concurrent-chunks",
            "4",
            "--max-concurrent-file-types",
            "3",
            "--token-expiry-retries",
            "2",
            "--transient-chunk-retries",
            "0",
            "--rate-limit-retries",
            "3",
            "--rate-limit-backoff-secs",
            "0.5",
        ]
    )

    assert exit_code == 0
    assert captured["max_concurrent_chunks"] == 4
    assert captured["max_concurrent_file_types"] == 3
    assert captured["token_expiry_retries"] == 2
    assert captured["transient_chunk_retries"] == 0
    assert captured["rate_limit_retries"] == 3
    assert captured["rate_limit_backoff_secs"] == 0.5
    assert captured["fixture_registry"] == str(registry_path)
